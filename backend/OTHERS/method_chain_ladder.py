from __future__ import annotations

from pathlib import Path
import json

import openpyxl
import pandas as pd

from common_triangle import build_chain_ladder_data, find_ibnr_workbook


def main() -> None:
    root = Path(__file__).resolve().parents[2]
    workbook = find_ibnr_workbook(root)
    out_dir = Path(__file__).resolve().parent / "outputs"
    out_dir.mkdir(parents=True, exist_ok=True)

    data = build_chain_ladder_data(workbook)

    # Excel cached benchmark from `calcule IBNR`.
    wb = openpyxl.load_workbook(workbook, data_only=True)
    ws = wb["calcule IBNR"]
    excel_factors = {1: ws.cell(29, 3).value, 2: ws.cell(29, 4).value, 3: ws.cell(29, 5).value}
    excel_ibnr = {
        2022: ws.cell(33, 9).value,
        2023: ws.cell(34, 9).value,
        2024: ws.cell(35, 9).value,
        2025: ws.cell(36, 9).value,
    }
    excel_total = ws.cell(37, 9).value

    py_total = sum(data.ibnr.values())

    cmp = pd.DataFrame(
        {
            "AY": [2022, 2023, 2024, 2025],
            "IBNR_python": [data.ibnr[a] for a in [2022, 2023, 2024, 2025]],
            "IBNR_excel": [excel_ibnr[a] for a in [2022, 2023, 2024, 2025]],
        }
    )
    cmp["diff"] = cmp["IBNR_python"] - cmp["IBNR_excel"]
    cmp.to_csv(out_dir / "chain_ladder_compare.csv", index=False)

    payload = {
        "method": "Chain Ladder (weighted factors)",
        "workbook": str(workbook),
        "factors_python": data.factors,
        "factors_excel": excel_factors,
        "total_ibnr_python": py_total,
        "total_ibnr_excel": excel_total,
        "total_diff": py_total - excel_total,
        "max_abs_ay_diff": float(cmp["diff"].abs().max()),
    }
    with open(out_dir / "chain_ladder_summary.json", "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)

    print("[Chain Ladder] Factors python:", data.factors)
    print("[Chain Ladder] Factors excel :", excel_factors)
    print("[Chain Ladder] IBNR total python:", f"{py_total:,.2f}")
    print("[Chain Ladder] IBNR total excel :", f"{excel_total:,.2f}")
    print("[Chain Ladder] total diff:", f"{(py_total-excel_total):+.6f}")


if __name__ == "__main__":
    main()
