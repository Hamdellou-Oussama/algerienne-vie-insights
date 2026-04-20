"""Workbook inventory extraction and markdown reporting."""

from __future__ import annotations

from hashlib import sha256
import logging
from pathlib import Path

from openpyxl import load_workbook

from src.domain.types import InventorySheet, WorkbookInventory
from src.preprocessing.schema_registry import get_dataset_contracts

LOGGER = logging.getLogger(__name__)


def _file_sha256(path: Path) -> str:
    """Compute the SHA-256 hash of a file."""

    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def inventory_workbook(path: Path, header_row_by_sheet: dict[str, int]) -> WorkbookInventory:
    """Extract workbook inventory including sheet dimensions and headers."""

    workbook = None
    try:
        workbook = load_workbook(path, data_only=False, read_only=False)
        sheets: list[InventorySheet] = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            header_row = header_row_by_sheet.get(sheet_name, 1)
            header_values = tuple(worksheet.cell(header_row, column).value for column in range(1, worksheet.max_column + 1))
            sheets.append(
                InventorySheet(
                    sheet_name=sheet_name,
                    row_count=worksheet.max_row,
                    column_count=worksheet.max_column,
                    merged_ranges=tuple(str(item) for item in worksheet.merged_cells.ranges),
                    header_row=header_row,
                    header_values=header_values,
                )
            )
        return WorkbookInventory(
            workbook_name=path.name,
            workbook_path=path,
            sha256=_file_sha256(path),
            sheets=tuple(sheets),
        )
    except Exception as exc:
        LOGGER.error("Failed to inventory workbook %s: %s", path, exc)
        raise
    finally:
        if workbook is not None:
            workbook.close()


def inventory_all_workbooks() -> list[WorkbookInventory]:
    """Inventory all known workbooks from the dataset registry."""

    inventories: list[WorkbookInventory] = []
    by_workbook: dict[Path, dict[str, int]] = {}
    for contract in get_dataset_contracts().values():
        sheet_map = by_workbook.setdefault(contract.workbook_path, {})
        sheet_map[contract.sheet_name] = contract.header_row
    for workbook_path, header_rows in by_workbook.items():
        inventories.append(inventory_workbook(workbook_path, header_rows))
    return inventories


def write_inventory_report(output_path: Path, inventories: list[WorkbookInventory]) -> Path:
    """Write workbook inventory to markdown."""

    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        lines = ["# Raw Workbook Inventory", ""]
        for inventory in inventories:
            lines.append(f"## {inventory.workbook_name}")
            lines.append(f"- path: `{inventory.workbook_path}`")
            lines.append(f"- sha256: `{inventory.sha256}`")
            lines.append("")
            for sheet in inventory.sheets:
                lines.append(f"### {sheet.sheet_name}")
                lines.append(f"- rows: {sheet.row_count}")
                lines.append(f"- columns: {sheet.column_count}")
                lines.append(f"- header_row: {sheet.header_row}")
                lines.append(f"- merged_ranges: `{list(sheet.merged_ranges)}`")
                lines.append(f"- header_values: `{list(sheet.header_values)}`")
                lines.append("")
        output_path.write_text("\n".join(lines), encoding="utf-8")
        LOGGER.info("Wrote workbook inventory to %s", output_path)
        return output_path
    except Exception as exc:
        LOGGER.error("Failed to write inventory report: %s", exc)
        raise
