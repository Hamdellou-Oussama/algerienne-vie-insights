"""Run the consolidated preprocessing, provision, reconciliation, and assumptions audit package."""

from __future__ import annotations

import argparse
from datetime import date, datetime
import json
import logging
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from src.orchestration.run_preprocessing import configure_logging, run_preprocessing
from src.preprocessing.ibnr_loader import IBNRLoader
from src.preprocessing.pb_loader import PBLoader
from src.preprocessing.pe_loader import PELoader
from src.preprocessing.ppna_loader import PPNALoader
from src.preprocessing.sap_loader import SAPLoader
from src.provisions.ibnr import calculate_ibnr
from src.provisions.ibnr_benktander import calculate_benktander
from src.provisions.ibnr_bf import calculate_bf
from src.provisions.ibnr_bootstrap import calculate_bootstrap
from src.provisions.ibnr_comparison import build_method_comparison
from src.provisions.ibnr_mack import calculate_mack
from src.provisions.pb import calculate_pb
from src.provisions.pe import calculate_pe
from src.provisions.ppna import calculate_ppna
from src.provisions.sap import calculate_sap
from src.reporting.assumptions import generate_assumption_registry
from src.reporting.reconciliation import run_reconciliation

LOGGER = logging.getLogger(__name__)
ROOT = Path(__file__).resolve().parents[2]


def _parse_ppna_sample_closing_date() -> date:
    """Read the PPNA sample closing date from the workbook production sheet."""

    workbook = load_workbook(ROOT / "data" / "level 01-level2-ÉCHANTILLON DATA PPNA.xlsx", data_only=True, read_only=True)
    value = workbook[" PRODUCTION"]["P1"].value
    if not isinstance(value, datetime):
        raise ValueError('PPNA workbook " PRODUCTION"!P1 did not contain a datetime closing date.')
    return value.date()


def _parse_sap_sample_closing_date() -> date:
    """Read the SAP sample closing date from the mentor-aligned workbook cell AC2."""

    workbook = load_workbook(ROOT / "data" / "level 01-DATA SAP groupe.xlsx", data_only=True, read_only=True)
    value = workbook["SAP GROUPE (2)"]["AC2"].value
    if not isinstance(value, datetime):
        raise ValueError("SAP workbook AC2 did not contain a datetime closing date.")
    return value.date()


def _write_markdown(summary: dict[str, Any], markdown_path: Path) -> None:
    """Write consolidated audit summary to Markdown."""

    lines = ["# Consolidated Audit Summary", ""]
    lines.append("## Provision Totals")
    for module_name, payload in summary["provision_totals"].items():
        lines.append(f"- {module_name}: `{payload}`")
    lines.append("")
    lines.append("## Blockers")
    for blocker in summary["blockers_summary"]:
        lines.append(f"- {blocker}")
    lines.append("")
    lines.append("## Artifacts")
    for key, value in summary["artifact_paths"].items():
        lines.append(f"- {key}: `{value}`")
    lines.append("")
    markdown_path.write_text("\n".join(lines), encoding="utf-8")


def run_audit(output_root: Path | None = None) -> dict[str, Any]:
    """Run preprocessing, provisions, assumptions, reconciliation, and write a consolidated package."""

    try:
        docs_root = output_root if output_root is not None else ROOT / "docs"
        validation_dir = docs_root / "validation_reports"
        validation_dir.mkdir(parents=True, exist_ok=True)

        preprocessing_summary = run_preprocessing(output_root=docs_root)

        ppna_result = calculate_ppna(PPNALoader().load().rows, closing_date=_parse_ppna_sample_closing_date())
        pe_result = calculate_pe(PELoader().load().rows)
        pb_result = calculate_pb(PBLoader().load().rows)
        sap_result = calculate_sap(SAPLoader().load().rows, closing_date=_parse_sap_sample_closing_date())
        from src.config import load_yaml_config
        ibnr_config = load_yaml_config("src/config/legislative.yaml")["ibnr"]
        ibnr_segment_by = "product" if ibnr_config.get("segment_by_product") else None
        ibnr_raw = calculate_ibnr(IBNRLoader().load().rows, segment_by=ibnr_segment_by)
        if isinstance(ibnr_raw, dict):
            ibnr_total = sum(r.total_ibnr for r in ibnr_raw.values())
            ibnr_closing_year = next(iter(ibnr_raw.values())).closing_year
            ibnr_shape = f"{len(next(iter(ibnr_raw.values())).by_occurrence_year)}x{next(iter(ibnr_raw.values())).max_development_year + 1}"
            ibnr_row_count = sum(sum(c.source_claim_count for c in r.triangle_cells if c.is_known) for r in ibnr_raw.values())
            ibnr_mode = "homogeneous_by_product"
            ibnr_segments = {k: v.total_ibnr for k, v in ibnr_raw.items()}
        else:
            ibnr_total = ibnr_raw.total_ibnr
            ibnr_closing_year = ibnr_raw.closing_year
            ibnr_shape = f"{len(ibnr_raw.by_occurrence_year)}x{ibnr_raw.max_development_year + 1}"
            ibnr_row_count = sum(c.source_claim_count for c in ibnr_raw.triangle_cells if c.is_known)
            ibnr_mode = "mixed_all_products"
            ibnr_segments = None

        # Secondary methods run for every IBNRResult regardless of segmentation mode.
        bf_cfg = ibnr_config.get("bornhuetter_ferguson", {})
        mack_cfg = ibnr_config.get("mack", {})
        bk_cfg = ibnr_config.get("benktander", {})
        boot_cfg = ibnr_config.get("bootstrap_odp", {})

        def _run_secondary(result: Any) -> dict[str, Any] | None:
            """Run all enabled secondary methods on one IBNRResult; return summary dict or None."""
            try:
                mack_r = calculate_mack(result) if mack_cfg.get("enabled", True) else None
                bf_r = calculate_bf(result, bf_cfg) if bf_cfg.get("enabled", True) else None
                bk_r = (
                    calculate_benktander(result, bf_r, bk_cfg)
                    if (bk_cfg.get("enabled", True) and bf_r is not None)
                    else None
                )
                boot_r = calculate_bootstrap(result, boot_cfg) if boot_cfg.get("enabled", True) else None
                if mack_r and bf_r and bk_r and boot_r:
                    cmp = build_method_comparison(result, mack_r, bf_r, bk_r, boot_r)
                    return {
                        "chain_ladder": result.total_ibnr,
                        "mack_se": mack_r.total_se_naive,
                        "bf_total": bf_r.total_ibnr_bf,
                        "benktander_total": bk_r.total_ibnr_benktander,
                        "bootstrap_mean": boot_r.mean_total_ibnr,
                        "bootstrap_p95": boot_r.percentiles.get(95),
                        "method_range": cmp.method_range,
                    }
            except Exception as exc:
                LOGGER.warning("Secondary IBNR methods failed (non-blocking): %s", exc)
            return None

        if isinstance(ibnr_raw, dict):
            # Homogeneous mode: run secondary methods per product segment.
            ibnr_method_comparison_by_product: dict[str, Any] = {}
            for product, seg_result in ibnr_raw.items():
                seg_cmp = _run_secondary(seg_result)
                if seg_cmp is not None:
                    ibnr_method_comparison_by_product[product] = seg_cmp
                    LOGGER.info("IBNR secondary methods done for product=%s range=%.2f", product, seg_cmp["method_range"])
            # Aggregate totals across products.
            ibnr_method_comparison: dict[str, Any] | None = (
                {
                    "chain_ladder": sum(v["chain_ladder"] for v in ibnr_method_comparison_by_product.values()),
                    "bf_total": sum(v["bf_total"] for v in ibnr_method_comparison_by_product.values()),
                    "benktander_total": sum(v["benktander_total"] for v in ibnr_method_comparison_by_product.values()),
                    "bootstrap_mean": sum(v["bootstrap_mean"] for v in ibnr_method_comparison_by_product.values()),
                    "by_product": ibnr_method_comparison_by_product,
                }
                if ibnr_method_comparison_by_product
                else None
            )
        else:
            # Mixed mode: single triangle, run once.
            ibnr_method_comparison = _run_secondary(ibnr_raw)
            if ibnr_method_comparison:
                LOGGER.info("IBNR method comparison complete: range=%.2f", ibnr_method_comparison["method_range"])

        provision_summary = {
            "ppna": {
                "closing_date": ppna_result.closing_date,
                "total_amount": ppna_result.total_amount,
                "row_count": len(ppna_result.row_results),
            },
            "pe": {
                "total_amount": pe_result.total_amount,
                "row_count": len(pe_result.row_results),
            },
            "pb": {
                "total_amount": pb_result.total_amount,
                "row_count": len(pb_result.row_results),
            },
            "sap": {
                "closing_date": sap_result.closing_date,
                "total_amount": sap_result.total_amount,
                "row_count": len(sap_result.row_results),
            },
            "ibnr": {
                "closing_year": ibnr_closing_year,
                "total_ibnr": ibnr_total,
                "triangle_shape": ibnr_shape,
                "row_count": ibnr_row_count,
                "mode": ibnr_mode,
                "by_product": ibnr_segments,
                "method_comparison": ibnr_method_comparison,
            },
        }
        provision_json_path = validation_dir / "provision_summary.json"
        provision_markdown_path = validation_dir / "provision_summary.md"
        provision_json_path.write_text(
            json.dumps(provision_summary, indent=2, ensure_ascii=False, sort_keys=True), encoding="utf-8"
        )
        provision_markdown_path.write_text(
            "# Provision Summary\n\n"
            + "\n".join(f"- {name}: `{payload}`" for name, payload in provision_summary.items()),
            encoding="utf-8",
        )

        assumptions = generate_assumption_registry(validation_dir)
        reconciliation = run_reconciliation(validation_dir)

        blockers_summary = []
        for entry in assumptions["payload"]["entries"]:
            if entry["status"] != "active":
                blockers_summary.append(f"assumption::{entry['assumption_name']}::{entry['status']}")
        for module in reconciliation["payload"]["modules"]:
            if module["status"] != "matched":
                blockers_summary.append(f"reconciliation::{module['module_name']}::{module['status']}")

        summary = {
            "preprocessing_summary": preprocessing_summary,
            "provision_totals": provision_summary,
            "reconciliation_summary": reconciliation["payload"]["status_counts"],
            "assumptions_summary": assumptions["payload"]["status_counts"],
            "blockers_summary": blockers_summary,
            "artifact_paths": {
                "preprocessing_summary": str(validation_dir / "preprocessing_summary.json"),
                "provision_summary_json": str(provision_json_path),
                "provision_summary_markdown": str(provision_markdown_path),
                "reconciliation_report_json": reconciliation["json_path"],
                "reconciliation_report_markdown": reconciliation["markdown_path"],
                "assumption_registry_json": assumptions["json_path"],
                "assumption_registry_markdown": assumptions["markdown_path"],
            },
        }
        summary_json_path = validation_dir / "audit_summary.json"
        summary_markdown_path = validation_dir / "audit_summary.md"
        summary_json_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False, sort_keys=True), encoding="utf-8")
        _write_markdown(summary, summary_markdown_path)
        LOGGER.info("Wrote consolidated audit summary to %s and %s", summary_json_path, summary_markdown_path)
        return summary
    except Exception as exc:
        LOGGER.error("Audit run failed: %s", exc)
        raise


def main() -> None:
    """CLI entry point for consolidated audit generation."""

    parser = argparse.ArgumentParser(description="Run the consolidated actuarial audit package.")
    parser.parse_args()
    configure_logging()
    summary = run_audit()
    print(json.dumps(summary, indent=2, ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()
