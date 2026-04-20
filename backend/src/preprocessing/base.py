"""Base dataset loader with lineage and anomaly capture."""

from __future__ import annotations

from collections import Counter
from copy import deepcopy
import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.domain.enums import DataType, RuleCategory, Severity
from src.domain.types import CleaningEvent, DatasetContract, DatasetResult, FieldContract
from src.preprocessing.lineage import build_lineage_record
from src.preprocessing.normalizers import normalize_category, normalize_date, normalize_decimal, normalize_integer
from src.preprocessing.validators import has_nonempty_values, is_allowed_category

LOGGER = logging.getLogger(__name__)


class BaseDatasetLoader:
    """Reusable workbook loader honoring schema contracts and full lineage."""

    def __init__(self, contract: DatasetContract) -> None:
        """Initialize a dataset loader."""

        self.contract = contract

    def load(self) -> DatasetResult:
        """Load, normalize, validate, and summarize a dataset."""

        formula_wb = None
        value_wb = None
        try:
            formula_wb = load_workbook(self.contract.workbook_path, data_only=False, read_only=True)
            value_wb = load_workbook(
                self.contract.workbook_path,
                data_only=self.contract.uses_data_only_values,
                read_only=True,
            )
            formula_ws = formula_wb[self.contract.sheet_name]
            value_ws = value_wb[self.contract.sheet_name]
            result = DatasetResult(dataset_name=self.contract.dataset_name)
            duplicate_counter: Counter[tuple[Any, ...]] = Counter()

            field_index = {field.raw_label: idx + 1 for idx, field in enumerate(self.contract.fields)}
            for row_number, (formula_row, value_row) in enumerate(
                zip(
                    formula_ws.iter_rows(min_row=self.contract.data_start_row, values_only=True),
                    value_ws.iter_rows(min_row=self.contract.data_start_row, values_only=True),
                ),
                start=self.contract.data_start_row,
            ):
                projected_values = [
                    formula_row[field_index[field.raw_label] - 1] if field_index[field.raw_label] - 1 < len(formula_row) else None
                    for field in self.contract.fields
                ]
                if not has_nonempty_values(projected_values):
                    continue

                canonical_row: dict[str, Any] = {"_source_row_number": row_number}
                raw_row: dict[str, Any] = {}
                for field in self.contract.fields:
                    index = field_index[field.raw_label] - 1
                    raw_value = formula_row[index] if index < len(formula_row) else None
                    value_for_normalization = value_row[index] if index < len(value_row) else raw_value
                    normalized_value, rule_applied = self._normalize_field(field, value_for_normalization)
                    severity = self._severity_for(field, raw_value, normalized_value, rule_applied)
                    canonical_row[field.canonical_name] = normalized_value
                    raw_row[field.canonical_name] = raw_value
                    result.lineage.append(
                        build_lineage_record(
                            dataset_name=self.contract.dataset_name,
                            workbook_name=self.contract.workbook_path.name,
                            sheet_name=self.contract.sheet_name,
                            row_number=row_number,
                            column_label=field.raw_label,
                            raw_value=raw_value,
                            normalized_value=normalized_value,
                            rule_applied=rule_applied,
                            severity=severity,
                        )
                    )
                    self._collect_field_events(result, field, row_number, raw_value, normalized_value, rule_applied)

                canonical_row["_raw_values"] = raw_row
                result.rows.append(canonical_row)

                duplicate_key = self._duplicate_key(canonical_row)
                if duplicate_key:
                    duplicate_counter[duplicate_key] += 1

            self._collect_duplicate_events(result, duplicate_counter)
            result.metrics = self._build_metrics(result)
            LOGGER.info("Loaded %s rows for dataset %s", len(result.rows), self.contract.dataset_name)
            return result
        except Exception as exc:
            LOGGER.error("Failed to load dataset %s: %s", self.contract.dataset_name, exc)
            raise
        finally:
            for workbook in (formula_wb, value_wb):
                if workbook is not None:
                    try:
                        workbook.close()
                    except Exception:
                        LOGGER.debug("Failed to close workbook handle for %s", self.contract.workbook_path)

    def _normalize_field(self, field: FieldContract, value: Any) -> tuple[Any, str]:
        """Normalize a single field according to its data type."""

        if field.data_type == DataType.DATE:
            return normalize_date(value)
        if field.data_type == DataType.DECIMAL:
            return normalize_decimal(value)
        if field.data_type == DataType.INTEGER:
            return normalize_integer(value)
        if field.data_type == DataType.CATEGORY:
            return normalize_category(value)
        if value in (None, ""):
            return None, "missing_string"
        return str(value).strip(), "string_strip"

    def _severity_for(self, field: FieldContract, raw_value: Any, normalized_value: Any, rule_applied: str) -> Severity:
        """Determine lineage severity for a normalized field."""

        if normalized_value is None and not field.nullable:
            return Severity.ERROR
        if rule_applied in {"string_to_iso", "excel_serial_to_iso", "string_to_decimal", "category_casefold"}:
            return Severity.INFO
        if rule_applied.startswith("ambiguous") or rule_applied.startswith("unsupported") or rule_applied == "invalid_decimal":
            return Severity.WARNING
        if raw_value != normalized_value:
            return Severity.INFO
        return Severity.DEBUG

    def _collect_field_events(
        self,
        result: DatasetResult,
        field: FieldContract,
        row_number: int,
        raw_value: Any,
        normalized_value: Any,
        rule_applied: str,
    ) -> None:
        """Emit structured cleaning events from field-level normalization and validation."""

        if normalized_value is None and raw_value in (None, "") and field.nullable:
            return
        if normalized_value is None and not field.nullable:
            result.events.append(
                CleaningEvent(
                    dataset_name=result.dataset_name,
                    row_number=row_number,
                    field_name=field.canonical_name,
                    category=RuleCategory.MANUAL_REVIEW_REQUIRED,
                    severity=Severity.ERROR,
                    message=f"Missing or unparseable required field: {field.raw_label}",
                    raw_value=raw_value,
                    normalized_value=normalized_value,
                )
            )
            return
        if field.data_type == DataType.DATE and isinstance(raw_value, str):
            result.events.append(
                CleaningEvent(
                    dataset_name=result.dataset_name,
                    row_number=row_number,
                    field_name=field.canonical_name,
                    category=RuleCategory.DATE_NORMALIZATION,
                    severity=Severity.WARNING,
                    message="Date arrived as string and required normalization.",
                    raw_value=raw_value,
                    normalized_value=normalized_value,
                )
            )
        if field.data_type in {DataType.DECIMAL, DataType.INTEGER} and isinstance(raw_value, str):
            result.events.append(
                CleaningEvent(
                    dataset_name=result.dataset_name,
                    row_number=row_number,
                    field_name=field.canonical_name,
                    category=RuleCategory.TYPE_COERCION,
                    severity=Severity.WARNING,
                    message="Numeric field arrived as string and required coercion.",
                    raw_value=raw_value,
                    normalized_value=normalized_value,
                )
            )
        if field.data_type == DataType.CATEGORY and raw_value not in (None, "") and str(raw_value).strip().casefold() != normalized_value:
            result.events.append(
                CleaningEvent(
                    dataset_name=result.dataset_name,
                    row_number=row_number,
                    field_name=field.canonical_name,
                    category=RuleCategory.CATEGORICAL_NORMALIZATION,
                    severity=Severity.INFO,
                    message="Category normalized for case and spacing.",
                    raw_value=raw_value,
                    normalized_value=normalized_value,
                )
            )
        if field.controlled_vocabulary and normalized_value is not None and not is_allowed_category(normalized_value, field.controlled_vocabulary):
            result.events.append(
                CleaningEvent(
                    dataset_name=result.dataset_name,
                    row_number=row_number,
                    field_name=field.canonical_name,
                    category=RuleCategory.BUSINESS_RULE_VALIDATION,
                    severity=Severity.WARNING,
                    message=f"Category {normalized_value!r} not in controlled vocabulary.",
                    raw_value=raw_value,
                    normalized_value=normalized_value,
                )
            )

    def _duplicate_key(self, row: dict[str, Any]) -> tuple[Any, ...] | None:
        """Return the dataset-specific duplicate key or None."""

        return None

    def _collect_duplicate_events(self, result: DatasetResult, duplicate_counter: Counter[tuple[Any, ...]]) -> None:
        """Emit duplicate detection events."""

        for key, count in duplicate_counter.items():
            if count > 1:
                result.events.append(
                    CleaningEvent(
                        dataset_name=result.dataset_name,
                        row_number=-1,
                        field_name="duplicate_key",
                        category=RuleCategory.DUPLICATE_DETECTION,
                        severity=Severity.WARNING,
                        message=f"Duplicate business key observed {count} times: {key!r}",
                        raw_value=key,
                        normalized_value=count,
                    )
                )

    def _build_metrics(self, result: DatasetResult) -> dict[str, Any]:
        """Return dataset metrics used by regression tests and reporting."""

        return {
            "row_count": len(result.rows),
            "event_count": len(result.events),
            "lineage_count": len(result.lineage),
        }

    def clone_row_without_raw_values(self, row: dict[str, Any]) -> dict[str, Any]:
        """Return a copy of a canonical row excluding raw value baggage."""

        row_copy = deepcopy(row)
        row_copy.pop("_raw_values", None)
        return row_copy
