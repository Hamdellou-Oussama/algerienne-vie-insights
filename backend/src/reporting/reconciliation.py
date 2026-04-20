"""Excel formula reconciliation against Python provision engines."""

from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import date, datetime
import json
import logging
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from src.preprocessing.ibnr_loader import IBNRLoader
from src.preprocessing.pe_loader import PELoader
from src.preprocessing.ppna_loader import PPNALoader
from src.preprocessing.sap_loader import SAPLoader
from src.provisions.ibnr import calculate_ibnr
from src.provisions.pe import calculate_pe
from src.provisions.ppna import calculate_ppna
from src.provisions.sap import calculate_sap

LOGGER = logging.getLogger(__name__)
ROOT = Path(__file__).resolve().parents[2]


@dataclass(frozen=True)
class ReconciliationModuleResult:
    """Reconciliation outcome for a single module."""

    module_name: str
    status: str
    python_total: float | None
    excel_total: float | None
    difference: float | None
    row_match_count: int
    row_mismatch_count: int
    notes: list[str]
    evidence: dict[str, Any]


@dataclass(frozen=True)
class ReconciliationReport:
    """Full reconciliation report across implemented provisions."""

    modules: list[ReconciliationModuleResult]


def _parse_output_sheet_date(text: str) -> date:
    """Parse a dd/mm/yyyy date embedded in a sheet label."""

    match = re.search(r"(\d{2}/\d{2}/\d{4})", text)
    if not match:
        raise ValueError(f"Could not find closing date in output label: {text!r}")
    return datetime.strptime(match.group(1), "%d/%m/%Y").date()


def _compare_float(left: float | None, right: float | None, tolerance: float = 1e-6) -> bool:
    """Compare two floats with tolerance."""

    if left is None or right is None:
        return False
    return abs(left - right) <= tolerance


def _reconcile_sap() -> ReconciliationModuleResult:
    """Reconcile SAP engine outputs against the mentor-aligned workbook formulas."""

    workbook_formula = load_workbook(ROOT / "data" / "level 01-DATA SAP groupe.xlsx", data_only=False, read_only=True)
    workbook_values = load_workbook(ROOT / "data" / "level 01-DATA SAP groupe.xlsx", data_only=True, read_only=True)
    closing_date_value = workbook_values["SAP GROUPE (2)"]["AC2"].value
    if not isinstance(closing_date_value, datetime):
        raise ValueError("SAP workbook AC2 did not contain a datetime closing date.")
    closing_date = closing_date_value.date()

    result = calculate_sap(SAPLoader().load().rows, closing_date=closing_date)
    worksheet_values = workbook_values["SAP GROUPE (2)"]
    output_formula = workbook_formula["SAP GROUPE (2)"]["AG2"].value
    output_cached = workbook_values["SAP GROUPE (2)"]["AG2"].value

    row_match_count = 0
    row_mismatch_count = 0
    blank_notification_rows = 0
    blank_notification_by_status: dict[str, int] = {}
    sample_mismatches: list[dict[str, Any]] = []
    by_row_number = {row.source_row_number: row for row in result.row_results}
    for source_row_number in range(3, 237):
        excel_value = float(worksheet_values[f"AC{source_row_number}"].value or 0.0)
        python_value = by_row_number[source_row_number].sap_amount
        if _compare_float(python_value, excel_value):
            row_match_count += 1
        else:
            row_mismatch_count += 1
            if len(sample_mismatches) < 10:
                sample_mismatches.append(
                    {"source_row_number": source_row_number, "python_value": python_value, "excel_value": excel_value}
                )
        if worksheet_values[f"Y{source_row_number}"].value is None:
            blank_notification_rows += 1
            status = str(worksheet_values[f"X{source_row_number}"].value or "").upper()
            blank_notification_by_status[status] = blank_notification_by_status.get(status, 0) + 1

    notes = [
        "Workbook formula logic in AC rows is used as the authoritative SAP reconciliation rule.",
        'The mentor-aligned rule is driven by "Date de déclaration", "Date de notification reglement /REJET", "Statut", and "Montant réglé".',
        "Legacy AB-column formulas remain in the workbook but are not the authoritative path for the corrected SAP logic.",
    ]
    status = "matched" if _compare_float(result.total_amount, float(output_cached)) and row_mismatch_count == 0 else "needs_review"
    return ReconciliationModuleResult(
        module_name="sap",
        status=status,
        python_total=result.total_amount,
        excel_total=float(output_cached),
        difference=result.total_amount - float(output_cached),
        row_match_count=row_match_count,
        row_mismatch_count=row_mismatch_count,
        notes=notes,
        evidence={
            "closing_date": closing_date.isoformat(),
            "output_formula_cell": "SAP GROUPE (2)!AG2",
            "output_formula": output_formula,
            "blank_notification_rows": blank_notification_rows,
            "blank_notification_by_status": blank_notification_by_status,
            "sample_row_formula": workbook_formula["SAP GROUPE (2)"]["AC3"].value,
            "legacy_row_formula": workbook_formula["SAP GROUPE (2)"]["AB3"].value,
            "sample_mismatches": sample_mismatches,
        },
    )


def _reconcile_pe() -> ReconciliationModuleResult:
    """Reconcile PE engine outputs against workbook formulas and cached values."""

    workbook_formula = load_workbook(ROOT / "data" / "level 01-ÉCHANTILLON DATA PE.xlsx", data_only=False, read_only=True)
    workbook_values = load_workbook(ROOT / "data" / "level 01-ÉCHANTILLON DATA PE.xlsx", data_only=True, read_only=True)
    worksheet_formula = workbook_formula["PE"]
    worksheet_values = workbook_values["PE"]

    result = calculate_pe(PELoader().load().rows)
    by_row_number = {row.source_row_number: row for row in result.row_results}
    row_match_count = 0
    row_mismatch_count = 0
    total_excel = 0.0
    sample_mismatches: list[dict[str, Any]] = []
    for source_row_number in range(4, 129):
        if not any(worksheet_values.cell(source_row_number, column).value not in (None, "") for column in range(1, 5)):
            continue
        excel_value = float(worksheet_values[f"U{source_row_number}"].value or 0.0)
        python_value = by_row_number[source_row_number].equalization_provision
        total_excel += excel_value
        if _compare_float(python_value, excel_value):
            row_match_count += 1
        else:
            row_mismatch_count += 1
            if len(sample_mismatches) < 10:
                sample_mismatches.append(
                    {"source_row_number": source_row_number, "python_value": python_value, "excel_value": excel_value}
                )

    status = "matched" if _compare_float(result.total_amount, total_excel) and row_mismatch_count == 0 else "needs_review"
    return ReconciliationModuleResult(
        module_name="pe",
        status=status,
        python_total=result.total_amount,
        excel_total=total_excel,
        difference=result.total_amount - total_excel,
        row_match_count=row_match_count,
        row_mismatch_count=row_mismatch_count,
        notes=[
            "PE is reconciled against formula-backed U column cached values.",
            "Workbook formula evidence is read directly from the PE sheet, not from the OBJECTIF PE template.",
        ],
        evidence={
            "formula_examples": {
                "H4": worksheet_formula["H4"].value,
                "M4": worksheet_formula["M4"].value,
                "N4": worksheet_formula["N4"].value,
                "T4": worksheet_formula["T4"].value,
                "U4": worksheet_formula["U4"].value,
            },
            "sample_mismatches": sample_mismatches,
        },
    )


def _reconcile_ppna() -> ReconciliationModuleResult:
    """Reconcile PPNA engine outputs against production-sheet workbook formulas."""

    workbook_formula = load_workbook(ROOT / "data" / "level 01-level2-ÉCHANTILLON DATA PPNA.xlsx", data_only=False, read_only=True)
    workbook_values = load_workbook(ROOT / "data" / "level 01-level2-ÉCHANTILLON DATA PPNA.xlsx", data_only=True, read_only=True)
    production_formula = workbook_formula[" PRODUCTION"]
    production_values = workbook_values[" PRODUCTION"]

    closing_date_value = production_values["P1"].value
    if not isinstance(closing_date_value, datetime):
        raise ValueError('PPNA workbook " PRODUCTION"!P1 did not contain a datetime closing date.')
    closing_date = closing_date_value.date()
    result = calculate_ppna(PPNALoader().load().rows, closing_date=closing_date)

    row_match_count = 0
    row_mismatch_count = 0
    sample_mismatches: list[dict[str, Any]] = []
    production_rows = production_values.iter_rows(min_row=2, max_row=21267, min_col=1, max_col=14, values_only=True)
    for audit_row, worksheet_row in zip(result.row_results, production_rows):
        source_row_number = audit_row.source_row_number
        if worksheet_row[0] is None:
            continue
        excel_value = float(worksheet_row[13] or 0.0)
        python_value = audit_row.ppna_amount
        if _compare_float(python_value, excel_value):
            row_match_count += 1
        else:
            row_mismatch_count += 1
            if len(sample_mismatches) < 10:
                sample_mismatches.append(
                    {"source_row_number": source_row_number, "python_value": python_value, "excel_value": excel_value}
                )

    excel_total = float(production_values["Q3"].value or 0.0)
    notes = [
        'PPNA is reconciled against production-sheet formulas in "nb de jours non aquise", "nb de jours contrat", "%", and "Prime non acquise".',
        'The "ETAT DE SORTIE" sheet remains a template and is not used as authoritative ground truth.',
    ]
    status = "matched" if _compare_float(result.total_amount, excel_total) and row_mismatch_count == 0 else "needs_review"
    return ReconciliationModuleResult(
        module_name="ppna",
        status=status,
        python_total=result.total_amount,
        excel_total=excel_total,
        difference=result.total_amount - excel_total,
        row_match_count=row_match_count,
        row_mismatch_count=row_mismatch_count,
        notes=notes,
        evidence={
            "closing_date_cell": " PRODUCTION!P1",
            "closing_date": closing_date.isoformat(),
            "output_formula_cell": ' PRODUCTION!Q3',
            "formula_examples": {
                "K2": production_formula["K2"].value,
                "L2": production_formula["L2"].value,
                "M2": production_formula["M2"].value,
                "N2": production_formula["N2"].value,
                "Q3": production_formula["Q3"].value,
            },
            "template_output_label": workbook_formula["ETAT DE SORTIE"]["B2"].value,
            "sample_mismatches": sample_mismatches,
        },
    )


def _reconcile_ibnr() -> ReconciliationModuleResult:
    """Reconcile IBNR Chain Ladder engine outputs against cached workbook values."""

    workbook_formula = load_workbook(ROOT / "data" / "level 02-ÉCHANTILLON DATA IBNR.xlsx", data_only=False, read_only=True)
    workbook_values = load_workbook(ROOT / "data" / "level 02-ÉCHANTILLON DATA IBNR.xlsx", data_only=True, read_only=True)
    ws_values = workbook_values["calcule IBNR"]
    ws_formula = workbook_formula["calcule IBNR"]

    result = calculate_ibnr(IBNRLoader().load().rows)
    occ_map = {r.occurrence_year: r for r in result.by_occurrence_year}
    factor_map = {f.development_year: f for f in result.development_factors}

    # Cells to reconcile: factors C29/D29/E29, per-row G/H/I 33..36, total I37
    # Workbook layout: row 29 = factors; rows 33..36 = completed triangle; I37 = total IBNR
    excel_f1 = float(ws_values["C29"].value or 0.0)
    excel_f2 = float(ws_values["D29"].value or 0.0)
    excel_f3 = float(ws_values["E29"].value or 0.0)
    excel_total = float(ws_values["I37"].value or 0.0)

    # Per-row values: rows 33..36 correspond to occ years 2022..2025
    excel_occ_rows = {
        2022: {"G": float(ws_values["G33"].value or 0.0), "H": float(ws_values["H33"].value or 0.0), "I": float(ws_values["I33"].value or 0.0)},
        2023: {"G": float(ws_values["G34"].value or 0.0), "H": float(ws_values["H34"].value or 0.0), "I": float(ws_values["I34"].value or 0.0)},
        2024: {"G": float(ws_values["G35"].value or 0.0), "H": float(ws_values["H35"].value or 0.0), "I": float(ws_values["I35"].value or 0.0)},
        2025: {"G": float(ws_values["G36"].value or 0.0), "H": float(ws_values["H36"].value or 0.0), "I": float(ws_values["I36"].value or 0.0)},
    }

    row_match_count = 0
    row_mismatch_count = 0
    sample_mismatches: list[dict[str, Any]] = []

    for occ_year, excel_vals in excel_occ_rows.items():
        python_row = occ_map.get(occ_year)
        if python_row is None:
            row_mismatch_count += 1
            sample_mismatches.append({"occ_year": occ_year, "reason": "missing_occurrence_year"})
            continue
        checks = [
            ("diagonal", python_row.diagonal_cumulative, excel_vals["G"]),
            ("ultimate", python_row.ultimate, excel_vals["H"]),
            ("reserve", python_row.reserve, excel_vals["I"]),
        ]
        for field_name, python_val, excel_val in checks:
            if _compare_float(python_val, excel_val):
                row_match_count += 1
            else:
                row_mismatch_count += 1
                if len(sample_mismatches) < 10:
                    sample_mismatches.append({
                        "occ_year": occ_year,
                        "field": field_name,
                        "python_value": python_val,
                        "excel_value": excel_val,
                    })

    total_matched = (
        _compare_float(result.total_ibnr, excel_total)
        and _compare_float(factor_map[1].factor, excel_f1)
        and _compare_float(factor_map[2].factor, excel_f2)
        and _compare_float(factor_map[3].factor, excel_f3)
        and row_mismatch_count == 0
    )
    status = "matched" if total_matched else "needs_review"

    return ReconciliationModuleResult(
        module_name="ibnr",
        status=status,
        python_total=result.total_ibnr,
        excel_total=excel_total,
        difference=result.total_ibnr - excel_total,
        row_match_count=row_match_count,
        row_mismatch_count=row_mismatch_count,
        notes=[
            "IBNR reconciled against calcule IBNR sheet cached values: factors C29:E29, reserves I33:I36, total I37.",
            "Single combined triangle over all ADE products matches workbook convention; product segmentation pending mentor guidance.",
        ],
        evidence={
            "closing_year": result.closing_year,
            "occurrence_year_window": list(result.occurrence_year_window),
            "excel_f1_cell": "calcule IBNR!C29",
            "excel_f2_cell": "calcule IBNR!D29",
            "excel_f3_cell": "calcule IBNR!E29",
            "excel_total_cell": "calcule IBNR!I37",
            "python_f1": factor_map[1].factor,
            "python_f2": factor_map[2].factor,
            "python_f3": factor_map[3].factor,
            "excel_f1": excel_f1,
            "excel_f2": excel_f2,
            "excel_f3": excel_f3,
            "formula_examples": {
                "C29": ws_formula["C29"].value,
                "D29": ws_formula["D29"].value,
                "E29": ws_formula["E29"].value,
                "E34": ws_formula["E34"].value,
                "C36": ws_formula["C36"].value,
                "I33": ws_formula["I33"].value,
                "I37": ws_formula["I37"].value,
            },
            "sample_mismatches": sample_mismatches,
        },
    )


def _write_markdown(report: ReconciliationReport, markdown_path: Path) -> None:
    """Write the reconciliation report as Markdown."""

    lines = ["# Reconciliation Report", ""]
    for module in report.modules:
        lines.append(f"## {module.module_name.upper()}")
        lines.append(f"- status: `{module.status}`")
        lines.append(f"- python_total: `{module.python_total}`")
        lines.append(f"- excel_total: `{module.excel_total}`")
        lines.append(f"- difference: `{module.difference}`")
        lines.append(f"- row_match_count: `{module.row_match_count}`")
        lines.append(f"- row_mismatch_count: `{module.row_mismatch_count}`")
        for note in module.notes:
            lines.append(f"- note: {note}")
        lines.append(f"- evidence: `{json.dumps(module.evidence, ensure_ascii=False, sort_keys=True)}`")
        lines.append("")
    markdown_path.write_text("\n".join(lines), encoding="utf-8")


def run_reconciliation(output_dir: Path) -> dict[str, object]:
    """Run provision reconciliation and write JSON/Markdown artifacts."""

    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        report = ReconciliationReport(modules=[_reconcile_ppna(), _reconcile_pe(), _reconcile_sap(), _reconcile_ibnr()])
        payload = {
            "modules": [asdict(module) for module in report.modules],
            "status_counts": {
                "matched": sum(1 for module in report.modules if module.status == "matched"),
                "needs_review": sum(1 for module in report.modules if module.status == "needs_review"),
            },
        }
        json_path = output_dir / "reconciliation_report.json"
        markdown_path = output_dir / "reconciliation_report.md"
        json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False, sort_keys=True), encoding="utf-8")
        _write_markdown(report, markdown_path)
        LOGGER.info("Wrote reconciliation report to %s and %s", json_path, markdown_path)
        return {"json_path": str(json_path), "markdown_path": str(markdown_path), "payload": payload}
    except Exception as exc:
        LOGGER.error("Failed to run reconciliation: %s", exc)
        raise
