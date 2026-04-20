from __future__ import annotations

from dataclasses import dataclass, asdict
from pathlib import Path
import json
import math

import pandas as pd
import openpyxl


@dataclass
class CheckResult:
    test_name: str
    row: int
    col: str
    rule: str
    expected: float
    actual: float
    abs_diff: float
    rel_diff: float
    ok: bool


def to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def compare(expected: float, actual: float, abs_tol: float = 1e-6, rel_tol: float = 1e-10) -> tuple[bool, float, float]:
    abs_diff = abs(expected - actual)
    denom = max(1.0, abs(expected))
    rel_diff = abs_diff / denom
    ok = abs_diff <= abs_tol or rel_diff <= rel_tol
    return ok, abs_diff, rel_diff


def add_check(bucket: list[CheckResult], test_name: str, row: int, col: str, rule: str, expected: float, actual: float) -> None:
    ok, abs_diff, rel_diff = compare(expected, actual)
    bucket.append(
        CheckResult(
            test_name=test_name,
            row=row,
            col=col,
            rule=rule,
            expected=expected,
            actual=actual,
            abs_diff=abs_diff,
            rel_diff=rel_diff,
            ok=ok,
        )
    )


def run_tests(xlsx_path: Path) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    wb_formula = openpyxl.load_workbook(xlsx_path, data_only=False)
    wb_value = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws_f = wb_formula["PE"]
    ws_v = wb_value["PE"]

    checks: list[CheckResult] = []

    # Core row-level formula validation across rows where U formulas exist.
    for r in range(4, 129):
        e = to_float(ws_v.cell(r, 5).value)
        f = to_float(ws_v.cell(r, 6).value)
        g = to_float(ws_v.cell(r, 7).value)
        i = to_float(ws_v.cell(r, 9).value)
        j = to_float(ws_v.cell(r, 10).value)
        k = to_float(ws_v.cell(r, 11).value)
        l = to_float(ws_v.cell(r, 12).value)
        p = to_float(ws_v.cell(r, 16).value)
        q = to_float(ws_v.cell(r, 17).value)
        rr = to_float(ws_v.cell(r, 18).value)
        s = to_float(ws_v.cell(r, 19).value)

        expected_h = e + f + g
        expected_m = i + j + k + l
        expected_n = expected_h - expected_m
        expected_o = k + i - g

        formula_t = ws_f.cell(r, 20).value
        if isinstance(formula_t, str) and "/S" in formula_t.upper():
            expected_t = (p + q + rr) / s if s != 0 else 0.0
            t_rule = "T = SUM(P:R)/S"
        else:
            expected_t = (p + q + rr) / 3.0
            t_rule = "T = AVERAGE(P:R)"

        expected_u = 0.0 if expected_n < 0 else min(0.72 * expected_n, 0.15 * expected_t)

        add_check(checks, "credit_total", r, "H", "H = SUM(E:G)", expected_h, to_float(ws_v.cell(r, 8).value))
        add_check(checks, "debit_total", r, "M", "M = SUM(I:L)", expected_m, to_float(ws_v.cell(r, 13).value))
        add_check(checks, "technical_result", r, "N", "N = H - M", expected_n, to_float(ws_v.cell(r, 14).value))
        add_check(checks, "claim_charge", r, "O", "O = K + I - G", expected_o, to_float(ws_v.cell(r, 15).value))
        add_check(checks, "moving_average", r, "T", t_rule, expected_t, to_float(ws_v.cell(r, 20).value))
        add_check(checks, "pb_formula", r, "U", "U = IF(N<0,0,MIN(72%*N,15%*T))", expected_u, to_float(ws_v.cell(r, 21).value))

    # Aggregate check.
    expected_u2 = sum(to_float(ws_v.cell(r, 21).value) for r in range(4, 129))
    actual_u2 = to_float(ws_v.cell(2, 21).value)
    add_check(checks, "pb_total", 2, "U", "U2 = SUM(U4:U128)", expected_u2, actual_u2)

    # Non-negativity test for PB line values.
    for r in range(4, 129):
        actual_u = to_float(ws_v.cell(r, 21).value)
        add_check(checks, "pb_non_negative", r, "U", "U >= 0", max(0.0, actual_u), actual_u)

    # Branch test: if N < 0 then U == 0.
    for r in range(4, 129):
        n_val = to_float(ws_v.cell(r, 14).value)
        u_val = to_float(ws_v.cell(r, 21).value)
        expected_branch = 0.0 if n_val < 0 else u_val
        add_check(checks, "pb_branch_if_n_negative", r, "U", "If N<0 then U=0", expected_branch, u_val)

    checks_df = pd.DataFrame([asdict(c) for c in checks])
    mismatches_df = checks_df.loc[~checks_df["ok"]].copy()

    test_counts = checks_df.groupby("test_name")["ok"].agg(["count", "sum"]).reset_index()
    test_counts["fail"] = test_counts["count"] - test_counts["sum"]
    test_counts.rename(columns={"sum": "pass"}, inplace=True)

    summary = {
        "workbook": str(xlsx_path),
        "total_checks": int(len(checks_df)),
        "total_mismatches": int(len(mismatches_df)),
        "overall_pass_rate": float((len(checks_df) - len(mismatches_df)) / max(1, len(checks_df))),
        "pb_total_expected_u2": float(expected_u2),
        "pb_total_actual_u2": float(actual_u2),
        "pb_total_diff": float(abs(expected_u2 - actual_u2)),
        "per_test": test_counts.to_dict(orient="records"),
    }

    return checks_df, mismatches_df, summary


def main() -> None:
    root = Path(__file__).resolve().parents[2]
    xlsx = root / "data" / "level 01-ÉCHANTILLON  DATA PE.xlsx"
    out_dir = Path(__file__).resolve().parent
    out_dir.mkdir(parents=True, exist_ok=True)

    checks_df, mismatches_df, summary = run_tests(xlsx)

    checks_path = out_dir / "pe_formula_test_checks.csv"
    mismatches_path = out_dir / "pe_formula_test_mismatches.csv"
    summary_path = out_dir / "pe_formula_test_summary.json"

    checks_df.to_csv(checks_path, index=False, encoding="utf-8-sig")
    mismatches_df.to_csv(mismatches_path, index=False, encoding="utf-8-sig")
    with summary_path.open("w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2, ensure_ascii=False)

    if summary["total_mismatches"] > 0:
        print("Validation status: FAILED")
    else:
        print("Validation status: PASSED")
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    print(f"Wrote: {checks_path}")
    print(f"Wrote: {mismatches_path}")
    print(f"Wrote: {summary_path}")


if __name__ == "__main__":
    main()
