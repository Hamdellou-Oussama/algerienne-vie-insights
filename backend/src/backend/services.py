"""Backend service layer for auth, storage, documents, runs, dashboard, and audit."""

from __future__ import annotations

from dataclasses import asdict, is_dataclass
from datetime import date, datetime
import csv
import io
import json
import logging
from pathlib import Path
import sqlite3
from typing import Any, Mapping
from uuid import uuid4

from openpyxl import load_workbook

from src.backend.database import get_connection
from src.backend.security import (
    expiry_after_days,
    expiry_after_minutes,
    generate_token,
    hash_password,
    hash_token,
    now_utc,
    utc_isoformat,
    verify_password,
)
from src.backend.settings import BackendSettings
from src.preprocessing.cleaning_report import write_cleaning_report
from src.preprocessing.ibnr_loader import IBNRLoader
from src.preprocessing.pb_loader import PBLoader
from src.preprocessing.pe_loader import PELoader
from src.preprocessing.ppna_loader import PPNALoader
from src.preprocessing.sap_loader import SAPLoader
from src.preprocessing.schema_registry import get_dataset_contract
from src.provisions.ibnr import IBNROccurrenceYearAudit, IBNRResult, calculate_ibnr
from src.provisions.ibnr_benktander import calculate_benktander
from src.provisions.ibnr_benktander import BenktanderResult
from src.provisions.ibnr_bf import calculate_bf
from src.provisions.ibnr_bf import BFResult
from src.provisions.ibnr_bootstrap import calculate_bootstrap
from src.provisions.ibnr_bootstrap import BootstrapResult
from src.provisions.ibnr_comparison import build_method_comparison
from src.provisions.ibnr_mack import MackResult
from src.provisions.ibnr_mack import calculate_mack
from src.provisions.pb import calculate_pb, calculate_pb_for_row
from src.provisions.pe import calculate_pe
from src.provisions.ppna import calculate_ppna
from src.provisions.sap import calculate_sap
from src.reporting.assumptions import generate_assumption_registry

LOGGER = logging.getLogger(__name__)

ALLOWED_DOMAINS = ("ppna", "sap", "pe", "pb", "ibnr")
ALLOWED_ROLES = {"ADMIN", "HR", "VIEWER"}
ALLOWED_STATUSES = {"ACTIVE", "SUSPENDED"}
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
IBNR_METHOD_ALIASES = {
    "chain_ladder": "chain_ladder",
    "chain_ladder_volume_weighted": "chain_ladder",
    "mack": "mack_chain_ladder",
    "mack_chain_ladder": "mack_chain_ladder",
    "bf": "bornhuetter_ferguson",
    "bornhuetter_ferguson": "bornhuetter_ferguson",
    "benktander": "benktander_k2",
    "benktander_k2": "benktander_k2",
    "bootstrap": "bootstrap_odp",
    "bootstrap_odp": "bootstrap_odp",
}


class BackendError(Exception):
    """Backend service error carrying an HTTP-like status code."""

    def __init__(self, message: str, status_code: int = 400) -> None:
        super().__init__(message)
        self.message = message
        self.status_code = status_code


def _dict(row: sqlite3.Row | None) -> dict[str, Any] | None:
    """Convert a SQLite row to a plain dictionary."""

    return None if row is None else dict(row)


def _require_domain(domain: str) -> str:
    """Validate a supported actuarial domain."""

    normalized = domain.lower()
    if normalized not in ALLOWED_DOMAINS:
        raise BackendError(f"Unsupported domain: {domain!r}", status_code=404)
    return normalized


def _json_dumps(payload: Any) -> str:
    """Serialize JSON using stable ordering."""

    return json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _serialize(value: Any) -> Any:
    """Convert dataclasses and nested objects into JSON-friendly structures."""

    if is_dataclass(value):
        return _serialize(asdict(value))
    if isinstance(value, dict):
        return {str(key): _serialize(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_serialize(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (date,)):
        return value.isoformat()
    return value


def _read_sha256(path: Path) -> str:
    """Compute a SHA-256 digest for a stored file."""

    import hashlib

    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _write_bytes(path: Path, payload: bytes) -> None:
    """Write bytes atomically to a storage path."""

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(payload)


def _write_json(path: Path, payload: Any) -> str:
    """Write JSON and return its SHA-256 digest."""

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False, sort_keys=True), encoding="utf-8")
    return _read_sha256(path)


def _meta(meta: Mapping[str, Any] | None) -> dict[str, str | None]:
    """Normalize request metadata for audit purposes."""

    return {
        "ip_address": None if meta is None else meta.get("ip_address"),
        "user_agent": None if meta is None else meta.get("user_agent"),
    }


def record_audit_event(
    connection: sqlite3.Connection,
    *,
    actor_user_id: str | None,
    action: str,
    target_type: str,
    target_id: str,
    payload: Mapping[str, Any],
    ip_address: str | None,
    user_agent: str | None,
) -> dict[str, Any]:
    """Append a tamper-evident audit event to the hash chain."""

    import hashlib

    previous_row = connection.execute(
        "SELECT event_hash FROM audit_event ORDER BY occurred_at DESC, event_id DESC LIMIT 1"
    ).fetchone()
    previous_hash = previous_row["event_hash"] if previous_row else None
    occurred_at = utc_isoformat()
    canonical_payload = {
        "actor_user_id": actor_user_id,
        "action": action,
        "target_type": target_type,
        "target_id": target_id,
        "occurred_at": occurred_at,
        "ip_address": ip_address,
        "user_agent": user_agent,
        "payload": _serialize(payload),
    }
    event_hash = hashlib.sha256((_json_dumps(canonical_payload) + (previous_hash or "")).encode("utf-8")).hexdigest()
    event_id = f"evt_{uuid4().hex}"
    connection.execute(
        """
        INSERT INTO audit_event (
            event_id, actor_user_id, action, target_type, target_id, occurred_at,
            ip_address, user_agent, payload_json, previous_event_hash, event_hash
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            event_id,
            actor_user_id,
            action,
            target_type,
            target_id,
            occurred_at,
            ip_address,
            user_agent,
            json.dumps(_serialize(payload), ensure_ascii=False, sort_keys=True),
            previous_hash,
            event_hash,
        ),
    )
    return {
        "event_id": event_id,
        "occurred_at": occurred_at,
        "event_hash": event_hash,
        "previous_event_hash": previous_hash,
    }


def _validate_role(role: str) -> str:
    """Validate a user role."""

    normalized = role.upper()
    if normalized not in ALLOWED_ROLES:
        raise BackendError(f"Unsupported role: {role!r}", status_code=422)
    return normalized


def _validate_status(status: str) -> str:
    """Validate a user status."""

    normalized = status.upper()
    if normalized not in ALLOWED_STATUSES:
        raise BackendError(f"Unsupported status: {status!r}", status_code=422)
    return normalized


def list_users(settings: BackendSettings) -> list[dict[str, Any]]:
    """Return all user accounts."""

    with get_connection(settings.db_path) as connection:
        rows = connection.execute(
            "SELECT user_id, username, role, status, created_at, created_by FROM user_account ORDER BY username"
        ).fetchall()
    return [_dict(row) for row in rows]


def _revoke_sessions_for_user(connection: sqlite3.Connection, user_id: str) -> None:
    """Revoke all active sessions for a user."""

    connection.execute(
        "UPDATE user_session SET revoked_at = ? WHERE user_id = ? AND revoked_at IS NULL",
        (utc_isoformat(), user_id),
    )


def bootstrap_admin(
    settings: BackendSettings,
    *,
    username: str,
    password: str,
    meta: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Create the first admin account when the database is empty."""

    if not username or not password:
        raise BackendError("Username and password are required.", status_code=422)
    request_meta = _meta(meta)
    with get_connection(settings.db_path) as connection:
        existing = connection.execute("SELECT COUNT(*) AS count FROM user_account").fetchone()
        if existing["count"] != 0:
            raise BackendError("Bootstrap is only allowed before the first account exists.", status_code=409)
        user_id = f"usr_{uuid4().hex}"
        connection.execute(
            """
            INSERT INTO user_account (user_id, username, password_hash, role, status, created_at, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (user_id, username, hash_password(password), "ADMIN", "ACTIVE", utc_isoformat(), "bootstrap"),
        )
        record_audit_event(
            connection,
            actor_user_id=user_id,
            action="bootstrap_admin",
            target_type="user_account",
            target_id=user_id,
            payload={"username": username, "role": "ADMIN"},
            ip_address=request_meta["ip_address"],
            user_agent=request_meta["user_agent"],
        )
        connection.commit()
    return login(settings, username=username, password=password, meta=meta)


def create_user(
    settings: BackendSettings,
    *,
    actor_user_id: str,
    username: str,
    password: str,
    role: str,
    status: str = "ACTIVE",
    meta: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Create a new user account."""

    if not username or not password:
        raise BackendError("Username and password are required.", status_code=422)
    role = _validate_role(role)
    status = _validate_status(status)
    request_meta = _meta(meta)
    with get_connection(settings.db_path) as connection:
        user_id = f"usr_{uuid4().hex}"
        try:
            connection.execute(
                """
                INSERT INTO user_account (user_id, username, password_hash, role, status, created_at, created_by)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (user_id, username, hash_password(password), role, status, utc_isoformat(), actor_user_id),
            )
        except sqlite3.IntegrityError as exc:
            raise BackendError(f"Username already exists: {username}", status_code=409) from exc
        record_audit_event(
            connection,
            actor_user_id=actor_user_id,
            action="create_user",
            target_type="user_account",
            target_id=user_id,
            payload={"username": username, "role": role, "status": status},
            ip_address=request_meta["ip_address"],
            user_agent=request_meta["user_agent"],
        )
        connection.commit()
    return {"user_id": user_id, "username": username, "role": role, "status": status}


def update_user_role(
    settings: BackendSettings,
    *,
    actor_user_id: str,
    user_id: str,
    role: str,
    meta: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Update a user's role."""

    role = _validate_role(role)
    request_meta = _meta(meta)
    with get_connection(settings.db_path) as connection:
        row = connection.execute("SELECT username FROM user_account WHERE user_id = ?", (user_id,)).fetchone()
        if row is None:
            raise BackendError(f"Unknown user: {user_id}", status_code=404)
        connection.execute("UPDATE user_account SET role = ? WHERE user_id = ?", (role, user_id))
        record_audit_event(
            connection,
            actor_user_id=actor_user_id,
            action="update_user_role",
            target_type="user_account",
            target_id=user_id,
            payload={"username": row["username"], "role": role},
            ip_address=request_meta["ip_address"],
            user_agent=request_meta["user_agent"],
        )
        connection.commit()
    return {"user_id": user_id, "role": role}


def update_user_status(
    settings: BackendSettings,
    *,
    actor_user_id: str,
    user_id: str,
    status: str,
    meta: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Update a user's status and revoke sessions when suspended."""

    status = _validate_status(status)
    request_meta = _meta(meta)
    with get_connection(settings.db_path) as connection:
        row = connection.execute("SELECT username FROM user_account WHERE user_id = ?", (user_id,)).fetchone()
        if row is None:
            raise BackendError(f"Unknown user: {user_id}", status_code=404)
        connection.execute("UPDATE user_account SET status = ? WHERE user_id = ?", (status, user_id))
        if status == "SUSPENDED":
            _revoke_sessions_for_user(connection, user_id)
        record_audit_event(
            connection,
            actor_user_id=actor_user_id,
            action="update_user_status",
            target_type="user_account",
            target_id=user_id,
            payload={"username": row["username"], "status": status},
            ip_address=request_meta["ip_address"],
            user_agent=request_meta["user_agent"],
        )
        connection.commit()
    return {"user_id": user_id, "status": status}


def login(
    settings: BackendSettings,
    *,
    username: str,
    password: str,
    meta: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Authenticate a user and create a new session."""

    request_meta = _meta(meta)
    with get_connection(settings.db_path) as connection:
        row = connection.execute(
            "SELECT user_id, username, password_hash, role, status FROM user_account WHERE username = ?",
            (username,),
        ).fetchone()
        if row is None or not verify_password(password, row["password_hash"]):
            raise BackendError("Invalid username or password.", status_code=401)
        if row["status"] != "ACTIVE":
            raise BackendError("Account is suspended.", status_code=403)
        access_token = generate_token()
        refresh_token = generate_token()
        session_id = f"ses_{uuid4().hex}"
        connection.execute(
            """
            INSERT INTO user_session (
                session_id, user_id, access_token_hash, refresh_token_hash, issued_at,
                access_expires_at, refresh_expires_at, revoked_at, ip_address, user_agent
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                session_id,
                row["user_id"],
                hash_token(access_token),
                hash_token(refresh_token),
                utc_isoformat(),
                expiry_after_minutes(settings.access_token_ttl_minutes),
                expiry_after_days(settings.refresh_token_ttl_days),
                None,
                request_meta["ip_address"],
                request_meta["user_agent"],
            ),
        )
        record_audit_event(
            connection,
            actor_user_id=row["user_id"],
            action="login",
            target_type="user_session",
            target_id=session_id,
            payload={"username": row["username"], "role": row["role"]},
            ip_address=request_meta["ip_address"],
            user_agent=request_meta["user_agent"],
        )
        connection.commit()
    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "user": {
            "user_id": row["user_id"],
            "username": row["username"],
            "role": row["role"],
            "status": row["status"],
        },
    }


def refresh_session(
    settings: BackendSettings,
    *,
    refresh_token: str,
    meta: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Rotate tokens using a valid refresh token."""

    request_meta = _meta(meta)
    now = utc_isoformat()
    with get_connection(settings.db_path) as connection:
        row = connection.execute(
            """
            SELECT s.session_id, s.user_id, u.username, u.role, u.status
            FROM user_session AS s
            JOIN user_account AS u ON u.user_id = s.user_id
            WHERE s.refresh_token_hash = ? AND s.revoked_at IS NULL AND s.refresh_expires_at > ?
            """,
            (hash_token(refresh_token), now),
        ).fetchone()
        if row is None:
            raise BackendError("Refresh token is invalid or expired.", status_code=401)
        if row["status"] != "ACTIVE":
            raise BackendError("Account is suspended.", status_code=403)
        access_token = generate_token()
        new_refresh_token = generate_token()
        connection.execute(
            """
            UPDATE user_session
            SET access_token_hash = ?, refresh_token_hash = ?, issued_at = ?, access_expires_at = ?,
                refresh_expires_at = ?, ip_address = ?, user_agent = ?
            WHERE session_id = ?
            """,
            (
                hash_token(access_token),
                hash_token(new_refresh_token),
                utc_isoformat(),
                expiry_after_minutes(settings.access_token_ttl_minutes),
                expiry_after_days(settings.refresh_token_ttl_days),
                request_meta["ip_address"],
                request_meta["user_agent"],
                row["session_id"],
            ),
        )
        record_audit_event(
            connection,
            actor_user_id=row["user_id"],
            action="refresh_session",
            target_type="user_session",
            target_id=row["session_id"],
            payload={"username": row["username"]},
            ip_address=request_meta["ip_address"],
            user_agent=request_meta["user_agent"],
        )
        connection.commit()
    return {
        "access_token": access_token,
        "refresh_token": new_refresh_token,
        "token_type": "bearer",
        "user": {
            "user_id": row["user_id"],
            "username": row["username"],
            "role": row["role"],
            "status": row["status"],
        },
    }


def get_session_user(settings: BackendSettings, access_token: str) -> dict[str, Any]:
    """Resolve an access token to an active user session."""

    now = utc_isoformat()
    with get_connection(settings.db_path) as connection:
        row = connection.execute(
            """
            SELECT s.session_id, s.user_id, u.username, u.role, u.status
            FROM user_session AS s
            JOIN user_account AS u ON u.user_id = s.user_id
            WHERE s.access_token_hash = ? AND s.revoked_at IS NULL AND s.access_expires_at > ?
            """,
            (hash_token(access_token), now),
        ).fetchone()
        if row is None:
            raise BackendError("Access token is invalid or expired.", status_code=401)
        if row["status"] != "ACTIVE":
            raise BackendError("Account is suspended.", status_code=403)
    return _dict(row)  # type: ignore[return-value]


def logout(
    settings: BackendSettings,
    *,
    access_token: str,
    meta: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Revoke the current session."""

    request_meta = _meta(meta)
    session = get_session_user(settings, access_token)
    with get_connection(settings.db_path) as connection:
        connection.execute(
            "UPDATE user_session SET revoked_at = ? WHERE session_id = ?",
            (utc_isoformat(), session["session_id"]),
        )
        record_audit_event(
            connection,
            actor_user_id=session["user_id"],
            action="logout",
            target_type="user_session",
            target_id=session["session_id"],
            payload={"username": session["username"]},
            ip_address=request_meta["ip_address"],
            user_agent=request_meta["user_agent"],
        )
        connection.commit()
    return {"session_id": session["session_id"], "status": "revoked"}


def _contract_for_domain(domain: str, workbook_path: Path | None = None):
    """Return the dataset contract for a domain."""

    return get_dataset_contract(_require_domain(domain), workbook_path=workbook_path)


def _validate_workbook_upload(
    settings: BackendSettings,
    domain: str,
    *,
    filename: str,
    content: bytes,
) -> None:
    """Validate an uploaded workbook before storing it."""

    if not filename.lower().endswith(".xlsx"):
        raise BackendError("Only .xlsx uploads are supported.", status_code=415)
    if len(content) == 0:
        raise BackendError("Upload body is empty.", status_code=422)
    if len(content) > settings.upload_max_bytes:
        raise BackendError("Uploaded workbook exceeds the configured size limit.", status_code=413)
    contract = _contract_for_domain(domain)
    workbook = None
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=contract.uses_data_only_values)
    except Exception as exc:
        raise BackendError(f"Uploaded workbook could not be opened: {exc}", status_code=422) from exc
    try:
        if contract.sheet_name not in workbook.sheetnames:
            raise BackendError(
                f'Workbook is missing required authoritative sheet "{contract.sheet_name}" for domain {domain}.',
                status_code=422,
            )
    finally:
        if workbook is not None:
            workbook.close()


def _export_authoritative_sheet(domain: str, workbook_path: Path, csv_path: Path, txt_path: Path) -> None:
    """Export the authoritative sheet to CSV and TXT derivatives."""

    contract = _contract_for_domain(domain, workbook_path=workbook_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=contract.uses_data_only_values)
    try:
        worksheet = workbook[contract.sheet_name]
        max_col = len(contract.fields)
        blank_run = 0
        blank_run_limit = 100

        csv_path.parent.mkdir(parents=True, exist_ok=True)
        txt_path.parent.mkdir(parents=True, exist_ok=True)
        with csv_path.open("w", encoding="utf-8", newline="") as csv_handle, txt_path.open("w", encoding="utf-8") as txt_handle:
            writer = csv.writer(csv_handle)
            for row_index, row in enumerate(
                worksheet.iter_rows(min_row=contract.header_row, max_col=max_col, values_only=True),
                start=contract.header_row,
            ):
                if row_index >= contract.data_start_row:
                    if all(value in (None, "") for value in row):
                        blank_run += 1
                        if blank_run >= blank_run_limit:
                            break
                        continue
                    blank_run = 0
                values = ["" if value is None else str(value) for value in row]
                writer.writerow(values)
                txt_handle.write("\t".join(values).rstrip() + "\n")
    finally:
        workbook.close()


def _document_downloads(domain: str, document_id: str) -> dict[str, str]:
    """Build download URLs for a document."""

    return {
        "xlsx": f"/api/v1/{domain}/documents/{document_id}/download/xlsx",
        "csv": f"/api/v1/{domain}/documents/{document_id}/download/csv",
        "txt": f"/api/v1/{domain}/documents/{document_id}/download/txt",
    }


def upload_document(
    settings: BackendSettings,
    *,
    domain: str,
    filename: str,
    content: bytes,
    actor_user_id: str,
    meta: Mapping[str, Any] | None = None,
    document_id: str | None = None,
) -> dict[str, Any]:
    """Store an uploaded workbook and its immediate derivatives."""

    domain = _require_domain(domain)
    _validate_workbook_upload(settings, domain, filename=filename, content=content)
    request_meta = _meta(meta)
    sha256 = _read_sha256_from_bytes(content)
    version_id = f"ver_{uuid4().hex}"
    if document_id is None:
        document_id = f"doc_{uuid4().hex}"
    raw_dir = settings.storage_root / "raw" / domain / document_id / version_id
    derived_dir = settings.storage_root / "derived" / domain / document_id / version_id
    xlsx_path = raw_dir / "original.xlsx"
    csv_path = derived_dir / "source.csv"
    txt_path = derived_dir / "source.txt"
    _write_bytes(xlsx_path, content)
    _export_authoritative_sheet(domain, xlsx_path, csv_path, txt_path)
    uploaded_at = utc_isoformat()

    with get_connection(settings.db_path) as connection:
        existing = connection.execute(
            "SELECT domain FROM document WHERE document_id = ?",
            (document_id,),
        ).fetchone()
        if existing is None:
            connection.execute(
                """
                INSERT INTO document (document_id, domain, original_filename, created_at, created_by, current_version_id, status)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (document_id, domain, filename, uploaded_at, actor_user_id, version_id, "ACTIVE"),
            )
        elif existing["domain"] != domain:
            raise BackendError("Existing document belongs to another domain.", status_code=409)
        else:
            connection.execute(
                "UPDATE document SET original_filename = ?, current_version_id = ?, status = ? WHERE document_id = ?",
                (filename, version_id, "ACTIVE", document_id),
            )
        connection.execute(
            """
            INSERT INTO document_version (
                version_id, document_id, sha256, mime_type, size_bytes,
                storage_path_xlsx, storage_path_csv, storage_path_txt, uploaded_at, uploaded_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                version_id,
                document_id,
                sha256,
                XLSX_MIME,
                len(content),
                str(xlsx_path),
                str(csv_path),
                str(txt_path),
                uploaded_at,
                actor_user_id,
            ),
        )
        record_audit_event(
            connection,
            actor_user_id=actor_user_id,
            action="upload_document",
            target_type="document_version",
            target_id=version_id,
            payload={
                "domain": domain,
                "document_id": document_id,
                "filename": filename,
                "sha256": sha256,
                "size_bytes": len(content),
            },
            ip_address=request_meta["ip_address"],
            user_agent=request_meta["user_agent"],
        )
        record_audit_event(
            connection,
            actor_user_id=actor_user_id,
            action="derive_document_formats",
            target_type="document_version",
            target_id=version_id,
            payload={"csv_path": str(csv_path), "txt_path": str(txt_path)},
            ip_address=request_meta["ip_address"],
            user_agent=request_meta["user_agent"],
        )
        connection.commit()
    return {
        "document_id": document_id,
        "version_id": version_id,
        "domain": domain,
        "original_filename": filename,
        "sha256": sha256,
        "downloads": _document_downloads(domain, document_id),
    }


def _read_sha256_from_bytes(content: bytes) -> str:
    """Compute a SHA-256 digest for an in-memory payload."""

    import hashlib

    return hashlib.sha256(content).hexdigest()


def list_documents(settings: BackendSettings, domain: str) -> list[dict[str, Any]]:
    """List documents for one domain."""

    domain = _require_domain(domain)
    with get_connection(settings.db_path) as connection:
        rows = connection.execute(
            """
            SELECT d.document_id, d.domain, d.original_filename, d.created_at, d.created_by, d.current_version_id, d.status,
                   v.sha256, v.uploaded_at, v.uploaded_by
            FROM document AS d
            LEFT JOIN document_version AS v ON v.version_id = d.current_version_id
            WHERE d.domain = ?
            ORDER BY d.created_at DESC
            """,
            (domain,),
        ).fetchall()
    return [
        {
            **_dict(row),
            "downloads": _document_downloads(domain, row["document_id"]),
        }
        for row in rows
    ]


def get_document(settings: BackendSettings, domain: str, document_id: str) -> dict[str, Any]:
    """Return one document and its current version metadata."""

    domain = _require_domain(domain)
    with get_connection(settings.db_path) as connection:
        row = connection.execute(
            """
            SELECT d.document_id, d.domain, d.original_filename, d.created_at, d.created_by, d.current_version_id, d.status,
                   v.sha256, v.uploaded_at, v.uploaded_by
            FROM document AS d
            LEFT JOIN document_version AS v ON v.version_id = d.current_version_id
            WHERE d.document_id = ? AND d.domain = ?
            """,
            (document_id, domain),
        ).fetchone()
        if row is None:
            raise BackendError(f"Unknown document: {document_id}", status_code=404)
    return {**_dict(row), "downloads": _document_downloads(domain, document_id)}  # type: ignore[arg-type]


def search_documents(settings: BackendSettings, query: str | None = None, domain: str | None = None) -> list[dict[str, Any]]:
    """Search document metadata by filename and optional domain."""

    like = f"%{query or ''}%"
    sql = """
        SELECT d.document_id, d.domain, d.original_filename, d.created_at, d.created_by, d.current_version_id, d.status
        FROM document AS d
        WHERE d.original_filename LIKE ?
    """
    params: list[Any] = [like]
    if domain:
        sql += " AND d.domain = ?"
        params.append(_require_domain(domain))
    sql += " ORDER BY d.created_at DESC"
    with get_connection(settings.db_path) as connection:
        rows = connection.execute(sql, tuple(params)).fetchall()
    return [_dict(row) for row in rows]


def list_document_versions(settings: BackendSettings, document_id: str) -> list[dict[str, Any]]:
    """List all immutable versions for one document."""

    with get_connection(settings.db_path) as connection:
        rows = connection.execute(
            """
            SELECT version_id, document_id, sha256, mime_type, size_bytes, storage_path_xlsx, storage_path_csv,
                   storage_path_txt, uploaded_at, uploaded_by
            FROM document_version
            WHERE document_id = ?
            ORDER BY uploaded_at DESC
            """,
            (document_id,),
        ).fetchall()
    if not rows:
        raise BackendError(f"Unknown document: {document_id}", status_code=404)
    return [_dict(row) for row in rows]


def resolve_document_download(
    settings: BackendSettings,
    *,
    domain: str,
    document_id: str,
    file_format: str,
    version_id: str | None = None,
) -> dict[str, Any]:
    """Resolve a document download to a stored filesystem path."""

    domain = _require_domain(domain)
    if file_format not in {"xlsx", "csv", "txt"}:
        raise BackendError(f"Unsupported download format: {file_format!r}", status_code=404)
    column = {
        "xlsx": "storage_path_xlsx",
        "csv": "storage_path_csv",
        "txt": "storage_path_txt",
    }[file_format]
    with get_connection(settings.db_path) as connection:
        if version_id is None:
            row = connection.execute(
                """
                SELECT v.version_id, v.sha256, v.size_bytes, v.storage_path_xlsx, v.storage_path_csv, v.storage_path_txt
                FROM document AS d
                JOIN document_version AS v ON v.version_id = d.current_version_id
                WHERE d.document_id = ? AND d.domain = ?
                """,
                (document_id, domain),
            ).fetchone()
        else:
            row = connection.execute(
                """
                SELECT v.version_id, v.sha256, v.size_bytes, v.storage_path_xlsx, v.storage_path_csv, v.storage_path_txt
                FROM document AS d
                JOIN document_version AS v ON v.document_id = d.document_id
                WHERE d.document_id = ? AND d.domain = ? AND v.version_id = ?
                """,
                (document_id, domain, version_id),
            ).fetchone()
        if row is None:
            raise BackendError(f"Unknown document/version combination: {document_id}", status_code=404)
    return {
        "version_id": row["version_id"],
        "path": Path(row[column]),
        "sha256": row["sha256"],
        "size_bytes": row["size_bytes"],
    }


def _loader_for_domain(domain: str, workbook_path: Path):
    """Return a loader instance for a stored workbook."""

    mapping = {
        "ppna": PPNALoader,
        "sap": SAPLoader,
        "pe": PELoader,
        "pb": PBLoader,
        "ibnr": IBNRLoader,
    }
    return mapping[domain](workbook_path=workbook_path)


def _coerce_bool(value: Any, *, default: bool) -> bool:
    """Normalize run parameters that can be provided as booleans or strings."""

    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"1", "true", "yes", "y", "on"}:
            return True
        if normalized in {"0", "false", "no", "n", "off"}:
            return False
    raise ValueError(f"Invalid boolean value: {value!r}")


def _normalize_selected_ibnr_method(parameters: Mapping[str, Any]) -> str:
    """Normalize selected IBNR method from run parameters."""

    raw_method = str(parameters.get("selected_method", "chain_ladder")).strip().lower()
    normalized = IBNR_METHOD_ALIASES.get(raw_method)
    if normalized is None:
        supported = ", ".join(sorted(set(IBNR_METHOD_ALIASES.values())))
        raise ValueError(
            f"Unsupported IBNR selected_method: {raw_method!r}. Supported values: {supported}."
        )
    return normalized


def _clone_ibnr_result_with_method(
    base_result: IBNRResult,
    *,
    method: str,
    by_occurrence_year: list[IBNROccurrenceYearAudit],
    total_ibnr: float,
    extra_parameters: Mapping[str, Any] | None = None,
) -> IBNRResult:
    """Clone an IBNRResult while replacing point estimates and method metadata."""

    cloned_parameters = dict(base_result.parameters)
    cloned_parameters["selected_method"] = method
    cloned_parameters["primary_method"] = method
    if extra_parameters:
        cloned_parameters.update(extra_parameters)
    return IBNRResult(
        closing_year=base_result.closing_year,
        occurrence_year_window=base_result.occurrence_year_window,
        max_development_year=base_result.max_development_year,
        method=method,
        total_ibnr=total_ibnr,
        triangle_cells=base_result.triangle_cells,
        development_factors=base_result.development_factors,
        by_occurrence_year=by_occurrence_year,
        parameters=cloned_parameters,
        excluded_rows=base_result.excluded_rows,
    )


def _bootstrap_primary_occurrence_rows(
    base_result: IBNRResult,
    bootstrap_result: BootstrapResult,
) -> tuple[list[IBNROccurrenceYearAudit], float]:
    """Allocate bootstrap mean IBNR to occurrence years using CL reserve weights."""

    target_total = float(bootstrap_result.mean_total_ibnr)
    base_rows = list(base_result.by_occurrence_year)
    if not base_rows:
        return [], target_total

    positive_weights = [max(float(row.reserve), 0.0) for row in base_rows]
    total_weight = sum(positive_weights)
    if total_weight <= 0.0:
        positive_weights = [1.0 for _ in base_rows]
        total_weight = float(len(base_rows))

    allocated_reserves = [target_total * (weight / total_weight) for weight in positive_weights]
    if allocated_reserves:
        allocated_reserves[-1] += target_total - sum(allocated_reserves)

    occurrence_rows: list[IBNROccurrenceYearAudit] = []
    for index, base_row in enumerate(base_rows):
        reserve = allocated_reserves[index]
        diagonal = float(base_row.diagonal_cumulative)
        occurrence_rows.append(
            IBNROccurrenceYearAudit(
                occurrence_year=base_row.occurrence_year,
                diagonal_cumulative=diagonal,
                ultimate=diagonal + reserve,
                reserve=reserve,
                last_known_development_year=base_row.last_known_development_year,
            )
        )

    return occurrence_rows, target_total


def _compute_ibnr_method_suite(
    base_result: IBNRResult,
) -> tuple[MackResult, BFResult, BenktanderResult, BootstrapResult, Any]:
    """Compute all secondary IBNR methods and the comparison summary from CL baseline."""

    mack_result = calculate_mack(base_result)
    bf_result = calculate_bf(base_result)
    benktander_result = calculate_benktander(base_result, bf_result)
    bootstrap_result = calculate_bootstrap(base_result)
    method_comparison = build_method_comparison(
        base_result,
        mack_result,
        bf_result,
        benktander_result,
        bootstrap_result,
    )
    return mack_result, bf_result, benktander_result, bootstrap_result, method_comparison


def _select_primary_ibnr_result(
    *,
    base_result: IBNRResult,
    selected_method: str,
    mack_result: MackResult,
    bf_result: BFResult,
    benktander_result: BenktanderResult,
    bootstrap_result: BootstrapResult,
) -> IBNRResult:
    """Build the primary IBNR result object from selected method output."""

    if selected_method == "chain_ladder":
        return _clone_ibnr_result_with_method(
            base_result,
            method="chain_ladder",
            by_occurrence_year=list(base_result.by_occurrence_year),
            total_ibnr=float(base_result.total_ibnr),
        )

    if selected_method == "mack_chain_ladder":
        return _clone_ibnr_result_with_method(
            base_result,
            method="mack_chain_ladder",
            by_occurrence_year=list(base_result.by_occurrence_year),
            total_ibnr=float(mack_result.total_ibnr),
            extra_parameters={"mack_total_se_naive": float(mack_result.total_se_naive)},
        )

    if selected_method == "bornhuetter_ferguson":
        bf_by_occurrence = {row.occurrence_year: row for row in bf_result.by_occurrence_year}
        occurrence_rows = [
            IBNROccurrenceYearAudit(
                occurrence_year=base_row.occurrence_year,
                diagonal_cumulative=float(base_row.diagonal_cumulative),
                ultimate=float(bf_by_occurrence[base_row.occurrence_year].ultimate_bf),
                reserve=float(bf_by_occurrence[base_row.occurrence_year].ibnr_bf),
                last_known_development_year=base_row.last_known_development_year,
            )
            for base_row in base_result.by_occurrence_year
        ]
        return _clone_ibnr_result_with_method(
            base_result,
            method="bornhuetter_ferguson",
            by_occurrence_year=occurrence_rows,
            total_ibnr=float(bf_result.total_ibnr_bf),
        )

    if selected_method == "benktander_k2":
        benktander_by_occurrence = {row.occurrence_year: row for row in benktander_result.by_occurrence_year}
        occurrence_rows = [
            IBNROccurrenceYearAudit(
                occurrence_year=base_row.occurrence_year,
                diagonal_cumulative=float(base_row.diagonal_cumulative),
                ultimate=float(benktander_by_occurrence[base_row.occurrence_year].ultimate_bk_k2),
                reserve=float(benktander_by_occurrence[base_row.occurrence_year].ibnr_bk_k2),
                last_known_development_year=base_row.last_known_development_year,
            )
            for base_row in base_result.by_occurrence_year
        ]
        return _clone_ibnr_result_with_method(
            base_result,
            method="benktander_k2",
            by_occurrence_year=occurrence_rows,
            total_ibnr=float(benktander_result.total_ibnr_benktander),
            extra_parameters={"benktander_k": int(benktander_result.k)},
        )

    if selected_method == "bootstrap_odp":
        occurrence_rows, total_ibnr = _bootstrap_primary_occurrence_rows(base_result, bootstrap_result)
        return _clone_ibnr_result_with_method(
            base_result,
            method="bootstrap_odp",
            by_occurrence_year=occurrence_rows,
            total_ibnr=total_ibnr,
            extra_parameters={
                "bootstrap_n_sim": int(bootstrap_result.n_sim),
                "bootstrap_random_seed": int(bootstrap_result.random_seed),
                "bootstrap_std_total_ibnr": float(bootstrap_result.std_total_ibnr),
                "bootstrap_percentiles": dict(bootstrap_result.percentiles),
            },
        )

    raise ValueError(f"Unsupported IBNR selected_method: {selected_method!r}")


def _run_calculation(
    domain: str,
    rows: list[dict[str, Any]],
    parameters: Mapping[str, Any],
) -> tuple[Any, Any, Any | None]:
    """Execute one actuarial module and return a result plus row-oriented payload."""

    if domain == "ppna":
        closing_date = date.fromisoformat(str(parameters["closing_date"]))
        result = calculate_ppna(rows, closing_date=closing_date)
        return result, result.row_results, None
    if domain == "sap":
        closing_date = date.fromisoformat(str(parameters["closing_date"]))
        result = calculate_sap(rows, closing_date=closing_date)
        return result, result.row_results, None
    if domain == "pe":
        result = calculate_pe(
            rows,
            positive_result_coefficient=parameters.get("positive_result_coefficient"),
            historical_average_coefficient=parameters.get("historical_average_coefficient"),
        )
        return result, result.row_results, None
    if domain == "pb":
        allow_override_value = parameters.get("allow_row_level_override")
        allow_override = (
            _coerce_bool(allow_override_value, default=True)
            if allow_override_value is not None
            else None
        )
        result = calculate_pb(
            rows,
            default_loss_ratio_threshold=parameters.get("default_loss_ratio_threshold"),
            default_pb_rate=parameters.get("default_pb_rate"),
            allow_row_level_override=allow_override,
        )
        return result, result.row_results, None
    if domain == "ibnr":
        selected_method = _normalize_selected_ibnr_method(parameters)
        segment_by = parameters.get("segment_by")
        if parameters.get("segment_by_product") and not segment_by:
            segment_by = "product"
        occurrence_year_window = parameters.get("occurrence_year_window")
        chain_ladder_result = calculate_ibnr(
            rows,
            closing_year=parameters.get("closing_year"),
            occurrence_year_window=tuple(occurrence_year_window) if occurrence_year_window else None,
            segment_by=segment_by,
        )
        if isinstance(chain_ladder_result, dict):
            primary_result: dict[str, IBNRResult] = {}
            comparison_payload: dict[str, Any] = {}
            for segment, base_segment_result in chain_ladder_result.items():
                mack_result, bf_result, benktander_result, bootstrap_result, method_comparison = _compute_ibnr_method_suite(
                    base_segment_result
                )
                primary_result[segment] = _select_primary_ibnr_result(
                    base_result=base_segment_result,
                    selected_method=selected_method,
                    mack_result=mack_result,
                    bf_result=bf_result,
                    benktander_result=benktander_result,
                    bootstrap_result=bootstrap_result,
                )
                comparison_payload[segment] = method_comparison
            row_payload = {key: value.by_occurrence_year for key, value in primary_result.items()}
            return primary_result, row_payload, comparison_payload
        else:
            mack_result, bf_result, benktander_result, bootstrap_result, method_comparison = _compute_ibnr_method_suite(
                chain_ladder_result
            )
            primary_result = _select_primary_ibnr_result(
                base_result=chain_ladder_result,
                selected_method=selected_method,
                mack_result=mack_result,
                bf_result=bf_result,
                benktander_result=benktander_result,
                bootstrap_result=bootstrap_result,
            )
            row_payload = primary_result.by_occurrence_year
            return primary_result, row_payload, method_comparison
    raise BackendError(f"Unsupported domain: {domain}", status_code=404)


def _infer_closing_date_from_workbook(domain: str, workbook_path: Path) -> str | None:
    """Infer a closing date from authoritative workbook cells for PPNA/SAP domains."""

    if domain == "ppna":
        sheet_name, cell = " PRODUCTION", "P1"
    elif domain == "sap":
        sheet_name, cell = "SAP GROUPE (2)", "AC2"
    else:
        return None

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return None
        value = workbook[sheet_name][cell].value
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return None
    finally:
        workbook.close()


def _attach_ibnr_method_comparison(result: Any, result_payload: Any) -> Any:
    """Attach full IBNR method-comparison data to the serialized result payload."""

    if not isinstance(result, IBNRResult) or not isinstance(result_payload, dict):
        return result_payload

    mack_result = calculate_mack(result)
    bf_result = calculate_bf(result)
    benktander_result = calculate_benktander(result, bf_result)
    bootstrap_result = calculate_bootstrap(result)
    method_comparison = build_method_comparison(result, mack_result, bf_result, benktander_result, bootstrap_result)
    result_payload["method_comparison"] = asdict(method_comparison)
    return result_payload


def _build_pb_recompute_audit(
    rows: list[dict[str, Any]],
    pb_result: Any,
) -> dict[str, Any]:
    """Recompute PB row-by-row and compare with stored run outputs for drift checks."""

    parameters = dict(getattr(pb_result, "parameters", {}) or {})
    threshold = float(parameters.get("default_loss_ratio_threshold", 0.0))
    rate = float(parameters.get("default_pb_rate", 0.0))
    allow_override = bool(parameters.get("allow_row_level_override", True))

    stored_by_row = {
        int(item.source_row_number): item
        for item in list(getattr(pb_result, "row_results", []) or [])
    }

    recompute_rows: list[dict[str, Any]] = []
    total_recomputed = 0.0
    max_abs_delta = 0.0
    mismatch_count = 0

    for row in rows:
        source_row_number = int(row.get("_source_row_number", -1))
        recomputed = calculate_pb_for_row(
            row,
            default_loss_ratio_threshold=threshold,
            default_pb_rate=rate,
            allow_row_level_override=allow_override,
        )
        total_recomputed += recomputed.participation_beneficiaire

        stored = stored_by_row.get(source_row_number)
        stored_amount = float(stored.participation_beneficiaire) if stored is not None else 0.0
        delta = float(recomputed.participation_beneficiaire) - stored_amount
        abs_delta = abs(delta)
        max_abs_delta = max(max_abs_delta, abs_delta)
        if abs_delta > 1e-9:
            mismatch_count += 1

        recompute_rows.append(
            {
                "source_row_number": source_row_number,
                "stored_participation_beneficiaire": stored_amount,
                "recomputed_participation_beneficiaire": float(recomputed.participation_beneficiaire),
                "delta": delta,
                "abs_delta": abs_delta,
                "effective_pb_rate": float(recomputed.effective_pb_rate),
                "effective_loss_ratio_threshold": float(recomputed.effective_loss_ratio_threshold),
                "eligibility_reason": recomputed.eligibility_reason,
                "zero_reason": recomputed.zero_reason,
            }
        )

    total_stored = float(getattr(pb_result, "total_amount", 0.0))
    total_delta = total_recomputed - total_stored
    return {
        "summary": {
            "row_count": len(recompute_rows),
            "allow_row_level_override": allow_override,
            "default_pb_rate": rate,
            "default_loss_ratio_threshold": threshold,
            "stored_total": total_stored,
            "recomputed_total": total_recomputed,
            "total_delta": total_delta,
            "max_abs_delta": max_abs_delta,
            "mismatch_count": mismatch_count,
            "within_tolerance": max_abs_delta <= 1e-9 and abs(total_delta) <= 1e-9,
        },
        "rows": recompute_rows,
    }


def _validate_pb_zero_rate_control(pb_result: Any) -> None:
    """Enforce PB control rule for disabled overrides with zero default rate."""

    parameters = dict(getattr(pb_result, "parameters", {}) or {})
    allow_override = bool(parameters.get("allow_row_level_override", True))
    default_pb_rate = float(parameters.get("default_pb_rate", 0.0))
    total_amount = float(getattr(pb_result, "total_amount", 0.0))

    if not allow_override and abs(default_pb_rate) <= 1e-12 and abs(total_amount) > 1e-6:
        raise ValueError(
            "PB control rule failed: allow_row_level_override=false and default_pb_rate=0 must yield total PB=0."
        )


def create_run(
    settings: BackendSettings,
    *,
    domain: str,
    actor_user_id: str,
    parameters: Mapping[str, Any],
    meta: Mapping[str, Any] | None = None,
    document_id: str | None = None,
    version_id: str | None = None,
) -> dict[str, Any]:
    """Execute a calculation run against a stored document version."""

    domain = _require_domain(domain)
    request_meta = _meta(meta)
    run_id = f"run_{uuid4().hex}"
    started_at = utc_isoformat()
    with get_connection(settings.db_path) as connection:
        if version_id is not None:
            version_row = connection.execute(
                """
                SELECT d.document_id, v.version_id, v.storage_path_xlsx
                FROM document_version AS v
                JOIN document AS d ON d.document_id = v.document_id
                WHERE v.version_id = ? AND d.domain = ?
                """,
                (version_id, domain),
            ).fetchone()
        elif document_id is not None:
            version_row = connection.execute(
                """
                SELECT d.document_id, v.version_id, v.storage_path_xlsx
                FROM document AS d
                JOIN document_version AS v ON v.version_id = d.current_version_id
                WHERE d.document_id = ? AND d.domain = ?
                """,
                (document_id, domain),
            ).fetchone()
        else:
            raise BackendError("Either document_id or version_id must be supplied.", status_code=422)
        if version_row is None:
            raise BackendError("Requested input document/version was not found.", status_code=404)

        resolved_parameters = dict(parameters)
        if domain in {"ppna", "sap"} and "closing_date" not in resolved_parameters:
            inferred_closing_date = _infer_closing_date_from_workbook(domain, Path(version_row["storage_path_xlsx"]))
            if inferred_closing_date is None:
                raise BackendError(
                    f"closing_date parameter is required for domain '{domain}' and could not be inferred from workbook.",
                    status_code=422,
                )
            resolved_parameters["closing_date"] = inferred_closing_date

        connection.execute(
            """
            INSERT INTO calculation_run (
                run_id, domain, document_version_id, parameters_json, status, started_at, finished_at,
                started_by, result_json, rows_json, error_message
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_id,
                domain,
                version_row["version_id"],
                json.dumps(_serialize(resolved_parameters), ensure_ascii=False, sort_keys=True),
                "running",
                started_at,
                None,
                actor_user_id,
                None,
                None,
                None,
            ),
        )
        record_audit_event(
            connection,
            actor_user_id=actor_user_id,
            action="start_calculation_run",
            target_type="calculation_run",
            target_id=run_id,
            payload={"domain": domain, "document_id": version_row["document_id"], "parameters": _serialize(resolved_parameters)},
            ip_address=request_meta["ip_address"],
            user_agent=request_meta["user_agent"],
        )
        connection.commit()

    try:
        workbook_path = Path(version_row["storage_path_xlsx"])
        dataset_result = _loader_for_domain(domain, workbook_path).load()
        output_dir = settings.storage_root / "runs" / domain / run_id
        cleaning_markdown_path, cleaning_json_path = write_cleaning_report({domain: dataset_result}, output_dir)
        result, row_payload, ibnr_method_comparison = _run_calculation(domain, dataset_result.rows, resolved_parameters)

        pb_audit_payload: dict[str, Any] | None = None
        if domain == "pb":
            _validate_pb_zero_rate_control(result)
            pb_audit_payload = _build_pb_recompute_audit(dataset_result.rows, result)

        result_payload = _serialize(result)
        if domain == "ibnr":
            if isinstance(result_payload, dict):
                if ibnr_method_comparison is not None:
                    result_payload["method_comparison"] = _serialize(ibnr_method_comparison)
                else:
                    result_payload = _attach_ibnr_method_comparison(result, result_payload)
        rows_payload = _serialize(row_payload)

        result_path = output_dir / "result.json"
        rows_path = output_dir / "rows.json"
        pb_audit_path = output_dir / "pb_audit.json"
        result_sha = _write_json(result_path, result_payload)
        rows_sha = _write_json(rows_path, rows_payload)
        cleaning_sha = _read_sha256(cleaning_json_path)
        pb_audit_sha = _write_json(pb_audit_path, pb_audit_payload) if pb_audit_payload is not None else None
        finished_at = utc_isoformat()

        with get_connection(settings.db_path) as connection:
            connection.execute(
                """
                UPDATE calculation_run
                SET status = ?, finished_at = ?, result_json = ?, rows_json = ?, error_message = ?
                WHERE run_id = ?
                """,
                (
                    "succeeded",
                    finished_at,
                    json.dumps(result_payload, ensure_ascii=False, sort_keys=True),
                    json.dumps(rows_payload, ensure_ascii=False, sort_keys=True),
                    None,
                    run_id,
                ),
            )
            artifacts = [
                ("result", "json", result_path, result_sha),
                ("rows", "json", rows_path, rows_sha),
                ("cleaning_report", "json", cleaning_json_path, cleaning_sha),
                ("cleaning_report", "md", cleaning_markdown_path, _read_sha256(cleaning_markdown_path)),
            ]
            if pb_audit_payload is not None and pb_audit_sha is not None:
                artifacts.append(("pb_audit", "json", pb_audit_path, pb_audit_sha))
            for artifact_type, artifact_format, storage_path, sha256 in artifacts:
                connection.execute(
                    """
                    INSERT INTO artifact (artifact_id, run_id, artifact_type, format, storage_path, sha256, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        f"art_{uuid4().hex}",
                        run_id,
                        artifact_type,
                        artifact_format,
                        str(storage_path),
                        sha256,
                        finished_at,
                    ),
                )
            record_audit_event(
                connection,
                actor_user_id=actor_user_id,
                action="finish_calculation_run",
                target_type="calculation_run",
                target_id=run_id,
                payload={"domain": domain, "result_path": str(result_path)},
                ip_address=request_meta["ip_address"],
                user_agent=request_meta["user_agent"],
            )
            connection.commit()
        return get_run(settings, domain, run_id)
    except Exception as exc:
        LOGGER.error("Calculation run %s failed: %s", run_id, exc)
        finished_at = utc_isoformat()
        with get_connection(settings.db_path) as connection:
            connection.execute(
                "UPDATE calculation_run SET status = ?, finished_at = ?, error_message = ? WHERE run_id = ?",
                ("failed", finished_at, str(exc), run_id),
            )
            record_audit_event(
                connection,
                actor_user_id=actor_user_id,
                action="fail_calculation_run",
                target_type="calculation_run",
                target_id=run_id,
                payload={"domain": domain, "error": str(exc)},
                ip_address=request_meta["ip_address"],
                user_agent=request_meta["user_agent"],
            )
            connection.commit()
        raise BackendError(f"Calculation run failed: {exc}", status_code=422) from exc


def _format_run(_: BackendSettings, domain: str, payload: dict[str, Any]) -> dict[str, Any]:
    """Normalize one calculation run payload for API responses."""

    run_id = str(payload["run_id"])
    payload["parameters"] = json.loads(payload.pop("parameters_json"))
    payload["artifacts"] = {
        "result": f"/api/v1/{domain}/runs/{run_id}/artifacts/result.json",
        "rows": f"/api/v1/{domain}/runs/{run_id}/artifacts/rows.json",
        "cleaning_report": f"/api/v1/{domain}/runs/{run_id}/artifacts/cleaning_report.json",
    }
    if domain == "pb":
        payload["artifacts"]["pb_audit"] = f"/api/v1/{domain}/runs/{run_id}/artifacts/pb_audit.json"
    return payload


def list_runs(
    settings: BackendSettings,
    domain: str,
    limit: int = 20,
    offset: int = 0,
    status: str | None = None,
) -> list[dict[str, Any]]:
    """List calculation runs for one domain in reverse start order."""

    domain = _require_domain(domain)
    with get_connection(settings.db_path) as connection:
        where = "WHERE domain = ?"
        params: list[Any] = [domain]
        if status:
            where += " AND status = ?"
            params.append(status)
        rows = connection.execute(
            f"""
            SELECT run_id, domain, document_version_id, parameters_json, status, started_at, finished_at, started_by, error_message
            FROM calculation_run
            {where}
            ORDER BY started_at DESC
            LIMIT ? OFFSET ?
            """,
            [*params, limit, offset],
        ).fetchall()
    return [_format_run(settings, domain, _dict(row) or {}) for row in rows]


def get_run(settings: BackendSettings, domain: str, run_id: str) -> dict[str, Any]:
    """Return one calculation run."""

    domain = _require_domain(domain)
    with get_connection(settings.db_path) as connection:
        row = connection.execute(
            """
            SELECT run_id, domain, document_version_id, parameters_json, status, started_at, finished_at, started_by, error_message
            FROM calculation_run
            WHERE run_id = ? AND domain = ?
            """,
            (run_id, domain),
        ).fetchone()
        if row is None:
            raise BackendError(f"Unknown run: {run_id}", status_code=404)
    return _format_run(settings, domain, _dict(row) or {})


def get_run_rows(settings: BackendSettings, domain: str, run_id: str) -> Any:
    """Return the row-level payload for a calculation run."""

    domain = _require_domain(domain)
    with get_connection(settings.db_path) as connection:
        row = connection.execute(
            "SELECT rows_json FROM calculation_run WHERE run_id = ? AND domain = ?",
            (run_id, domain),
        ).fetchone()
        if row is None:
            raise BackendError(f"Unknown run: {run_id}", status_code=404)
        if row["rows_json"] is None:
            raise BackendError("Run rows are not available for this run.", status_code=409)
    return json.loads(row["rows_json"])


def resolve_run_artifact(settings: BackendSettings, domain: str, run_id: str, artifact_name: str) -> dict[str, Any]:
    """Resolve a run artifact to a stored file path."""

    domain = _require_domain(domain)
    artifact_type, _, extension = artifact_name.partition(".")
    if not artifact_type or not extension:
        raise BackendError("Artifact names must include a file extension.", status_code=404)
    with get_connection(settings.db_path) as connection:
        row = connection.execute(
            """
            SELECT a.storage_path, a.sha256
            FROM artifact AS a
            JOIN calculation_run AS r ON r.run_id = a.run_id
            WHERE a.run_id = ? AND r.domain = ? AND a.artifact_type = ? AND a.format = ?
            """,
            (run_id, domain, artifact_type, extension),
        ).fetchone()
        if row is None:
            raise BackendError(f"Unknown artifact: {artifact_name}", status_code=404)
    return {"path": Path(row["storage_path"]), "sha256": row["sha256"]}


def _latest_successful_runs(settings: BackendSettings) -> dict[str, dict[str, Any]]:
    """Return the latest successful run per domain."""

    by_domain: dict[str, dict[str, Any]] = {}
    with get_connection(settings.db_path) as connection:
        rows = connection.execute(
            """
            SELECT run_id, domain, finished_at, result_json
            FROM calculation_run
            WHERE status = 'succeeded'
            ORDER BY finished_at DESC
            """
        ).fetchall()
    for row in rows:
        if row["domain"] not in by_domain:
            by_domain[row["domain"]] = _dict(row)  # type: ignore[assignment]
    return by_domain


def _extract_total(domain: str, result_payload: Any) -> float:
    """Extract the headline provision total from a run result."""

    if domain == "ibnr":
        if isinstance(result_payload, dict) and "total_ibnr" in result_payload:
            return float(result_payload["total_ibnr"])
        if isinstance(result_payload, dict):
            return float(sum(float(item["total_ibnr"]) for item in result_payload.values()))
        return 0.0
    return float(result_payload["total_amount"])


def build_dashboard_summary(settings: BackendSettings) -> dict[str, Any]:
    """Return current dashboard summary data."""

    latest_runs = _latest_successful_runs(settings)
    domain_totals: dict[str, Any] = {}
    grand_total = 0.0
    for domain, run in latest_runs.items():
        result_payload = json.loads(run["result_json"])
        total = _extract_total(domain, result_payload)
        domain_totals[domain] = {"run_id": run["run_id"], "finished_at": run["finished_at"], "total": total}
        grand_total += total
    return {
        "domains": domain_totals,
        "grand_total": grand_total,
        "completed_domains": len(domain_totals),
        "expected_domains": len(ALLOWED_DOMAINS),
    }


def build_dashboard_alerts(settings: BackendSettings) -> dict[str, Any]:
    """Return dashboard alerts driven by missing runs and open assumptions."""

    summary = build_dashboard_summary(settings)
    assumption_output = settings.storage_root / "dashboard_assumptions"
    assumption_payload = generate_assumption_registry(assumption_output)["payload"]
    missing_domains = [domain for domain in ALLOWED_DOMAINS if domain not in summary["domains"]]
    alerts: list[dict[str, Any]] = []
    for domain in missing_domains:
        alerts.append({"type": "missing_run", "domain": domain, "message": f"No successful run for {domain}."})
    if assumption_payload["status_counts"]["needs_review"]:
        alerts.append(
            {
                "type": "assumption_review",
                "count": assumption_payload["status_counts"]["needs_review"],
                "message": "Some assumptions still require mentor review.",
            }
        )
    return {"alerts": alerts}


def build_dashboard_timeline(settings: BackendSettings, limit: int = 25) -> dict[str, Any]:
    """Return recent audit events for the dashboard timeline."""

    with get_connection(settings.db_path) as connection:
        rows = connection.execute(
            """
            SELECT event_id, actor_user_id, action, target_type, target_id, occurred_at
            FROM audit_event
            ORDER BY occurred_at DESC, event_id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return {"events": [_dict(row) for row in rows]}


def build_dashboard_completion(settings: BackendSettings) -> dict[str, Any]:
    """Return simple completion status by domain."""

    summary = build_dashboard_summary(settings)
    return {
        "domains": {
            domain: {"completed": domain in summary["domains"]}
            for domain in ALLOWED_DOMAINS
        }
    }


def build_bilan_current(settings: BackendSettings) -> dict[str, Any]:
    """Return the current derived bilan snapshot."""

    latest_runs = _latest_successful_runs(settings)
    totals: dict[str, float] = {}
    source_runs: dict[str, str] = {}
    for domain, run in latest_runs.items():
        result_payload = json.loads(run["result_json"])
        totals[domain] = _extract_total(domain, result_payload)
        source_runs[domain] = run["run_id"]
    return {
        "generated_at": utc_isoformat(),
        "totals": totals,
        "grand_total": sum(totals.values()),
        "source_runs": source_runs,
    }


def create_bilan_snapshot(
    settings: BackendSettings,
    *,
    actor_user_id: str,
    meta: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Persist a derived bilan snapshot."""

    request_meta = _meta(meta)
    payload = build_bilan_current(settings)
    level3_payload = compute_level3_bilan(settings)
    snapshot_id = f"bil_{uuid4().hex}"
    created_at = utc_isoformat()
    with get_connection(settings.db_path) as connection:
        connection.execute(
            """
            INSERT INTO bilan_snapshot (snapshot_id, created_at, created_by, totals_json, source_runs_json, level3_json)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                snapshot_id,
                created_at,
                actor_user_id,
                json.dumps(payload["totals"], ensure_ascii=False, sort_keys=True),
                json.dumps(payload["source_runs"], ensure_ascii=False, sort_keys=True),
                json.dumps(level3_payload, ensure_ascii=False, sort_keys=True),
            ),
        )
        record_audit_event(
            connection,
            actor_user_id=actor_user_id,
            action="create_bilan_snapshot",
            target_type="bilan_snapshot",
            target_id=snapshot_id,
            payload=payload,
            ip_address=request_meta["ip_address"],
            user_agent=request_meta["user_agent"],
        )
        connection.commit()
    return {
        "snapshot_id": snapshot_id,
        "created_at": created_at,
        "created_by": actor_user_id,
        **payload,
        "level3": level3_payload,
    }


def list_bilan_history(settings: BackendSettings) -> list[dict[str, Any]]:
    """List historical bilan snapshots."""

    with get_connection(settings.db_path) as connection:
        rows = connection.execute(
            "SELECT snapshot_id, created_at, created_by, totals_json, source_runs_json, level3_json FROM bilan_snapshot ORDER BY created_at DESC"
        ).fetchall()
    history: list[dict[str, Any]] = []
    for row in rows:
        payload = _dict(row)
        totals = json.loads(payload.pop("totals_json"))
        source_runs = json.loads(payload.pop("source_runs_json"))
        level3_json = payload.pop("level3_json")
        payload["totals"] = totals
        payload["source_runs"] = source_runs
        payload["grand_total"] = float(sum(float(value) for value in totals.values()))
        payload["generated_at"] = payload["created_at"]
        payload["level3"] = json.loads(level3_json) if level3_json else None
        history.append(payload)
    return history


def compute_level3_bilan(settings: BackendSettings) -> dict[str, Any]:
    """Compute the Level3 bilan by year from the Level3 Bilan sinistres workbook.

    Implements the AC-logic SAP method (Method B) audited and validated in ./bilan.
    Rules sourced from bilan/rules.md:
      - En cours au 01/01/N  = dossiers/montant with SAP(ref=31/12/N-1) > 0
      - Declares en N        = declaration_year = N, declared_amount
      - Reglements en N      = status REGLE, notification_year = N, paid_amount
      - Rejets en N          = status REJET, notification_year = N, declared_amount
      - Repris en N          = status REPRIS, notification_year = N, declared_amount (zero in this dataset)
      - Reevaluation         = fixed to 0
      - Reserves au 31/12/N  = dossiers/montant with SAP(ref=31/12/N) > 0
      - Balance check: entrant + repris + declares - reglements - rejets + reeval = sortant
    """

    from datetime import date as _date

    from src.preprocessing.schema_registry import get_dataset_contract
    from src.provisions.sap import calculate_sap_for_row

    contract = get_dataset_contract("bilan_level3")
    if not contract.workbook_path.exists():
        raise BackendError(
            f"Level3 bilan workbook not found at {contract.workbook_path}. "
            "Please upload the workbook to the data directory.",
            status_code=404,
        )

    from src.preprocessing.sap_loader import SAPLoader

    loader = SAPLoader(workbook_path=contract.workbook_path)
    try:
        dataset = loader.load()
    except Exception as exc:
        raise BackendError(f"Failed to load Level3 bilan workbook: {exc}", status_code=500) from exc

    rows = dataset.rows

    # Determine exercise years from declaration_date
    decl_years: set[int] = set()
    for row in rows:
        dd = row.get("declaration_date")
        if dd:
            try:
                decl_years.add(int(str(dd)[:4]))
            except (ValueError, TypeError):
                pass
    if not decl_years:
        return {
            "generated_at": utc_isoformat(),
            "scope_mode": "portfolio",
            "years": [],
            "source": "level3_bilan_sinistres",
        }

    min_year = min(decl_years)
    max_year = max(decl_years) + 1  # include one carry-forward year

    def _parse_date(value: Any) -> _date | None:
        if value in (None, ""):
            return None
        try:
            return _date.fromisoformat(str(value)[:10])
        except ValueError:
            return None

    def _legacy_signed_sap_amount(row: Mapping[str, Any], ref: _date) -> float:
        """Historical SAP logic kept only for old-vs-new delta reporting."""

        declaration_date = _parse_date(row.get("declaration_date"))
        if declaration_date is None:
            return 0.0
        settlement_notification_date = _parse_date(row.get("settlement_notification_date"))
        status = str(row.get("status") or "").upper()
        declared_amount = float(row.get("declared_amount") or 0.0)
        paid_amount = float(row.get("paid_amount") or 0.0)

        if ref < declaration_date:
            return 0.0
        if settlement_notification_date is not None and declaration_date < ref < settlement_notification_date:
            return declared_amount
        if status == "REJET":
            return 0.0
        return declared_amount - paid_amount

    def _sap_at(ref_year: int, *, use_legacy_signed: bool = False) -> dict[str, float]:
        """Compute SAP per claim_id at year-end ref_year."""
        ref = _date(ref_year, 12, 31)
        sap_by_id: dict[str, float] = {}
        for row in rows:
            try:
                amount = (
                    _legacy_signed_sap_amount(row, ref)
                    if use_legacy_signed
                    else calculate_sap_for_row(row, ref).sap_amount
                )
                cid = str(row.get("claim_id") or row.get("_source_row_number"))
                sap_by_id[cid] = sap_by_id.get(cid, 0.0) + amount
            except Exception:
                pass
        return sap_by_id

    def _year_from_date(date_str: Any) -> int | None:
        if not date_str:
            return None
        try:
            return int(str(date_str)[:4])
        except (ValueError, TypeError):
            return None

    # Pre-compute SAP at each required reference date
    years = list(range(min_year, max_year + 1))
    sap_cache: dict[int, dict[str, float]] = {}
    sap_cache_legacy_signed: dict[int, dict[str, float]] = {}
    for yr in years:
        sap_cache[yr] = _sap_at(yr)
        sap_cache_legacy_signed[yr] = _sap_at(yr, use_legacy_signed=True)

    table: list[dict[str, Any]] = []
    prev_sortant_nbre = 0
    prev_sortant_montant = 0.0

    for idx, yr in enumerate(sorted(years)):
        sap_prev = sap_cache.get(yr - 1, {})
        sap_curr = sap_cache.get(yr, {})
        sap_curr_legacy_signed = sap_cache_legacy_signed.get(yr, {})

        # En cours (entrant) = SAP at 31/12/(N-1)
        if idx == 0:
            en_cours_nbre = sum(1 for v in sap_prev.values() if v > 0)
            en_cours_montant = sum(v for v in sap_prev.values() if v > 0)
        else:
            # Option A carry-forward: entrant(N) = sortant(N-1)
            en_cours_nbre = prev_sortant_nbre
            en_cours_montant = prev_sortant_montant

        # Declares en N: declaration_year = N
        dec_rows = [r for r in rows if _year_from_date(r.get("declaration_date")) == yr]
        declares_nbre = len(dec_rows)
        declares_montant = sum(float(r.get("declared_amount") or 0) for r in dec_rows)

        # Reglements en N: status REGLE and notification_year = N
        regl_rows = [
            r for r in rows
            if str(r.get("status") or "").upper() == "REGLE"
            and _year_from_date(r.get("settlement_notification_date")) == yr
        ]
        reglements_nbre = len(regl_rows)
        reglements_montant = sum(float(r.get("paid_amount") or 0) for r in regl_rows)

        # Rejets en N: status REJET and notification_year = N
        rej_rows = [
            r for r in rows
            if str(r.get("status") or "").upper() == "REJET"
            and _year_from_date(r.get("settlement_notification_date")) == yr
        ]
        rejet_nbre = len(rej_rows)
        rejet_montant = sum(float(r.get("declared_amount") or 0) for r in rej_rows)

        # Repris en N: status REPRIS and notification_year = N
        repris_rows = [
            r for r in rows
            if str(r.get("status") or "").upper() == "REPRIS"
            and _year_from_date(r.get("settlement_notification_date")) == yr
        ]
        repris_nbre = len(repris_rows)
        repris_montant = sum(float(r.get("declared_amount") or 0) for r in repris_rows)

        # Reevaluation: fixed to 0 for this dataset
        reevaluation_pos = 0.0
        reevaluation_neg = 0.0

        # Reserves (sortant): SAP at 31/12/N
        reserves_nbre = sum(1 for v in sap_curr.values() if v > 0)
        reserves_montant = sum(v for v in sap_curr.values() if v > 0)
        reserves_montant_old_signed = sum(v for v in sap_curr_legacy_signed.values() if v > 0)
        reserves_montant_delta_new_minus_old_signed = reserves_montant - reserves_montant_old_signed

        # Balance check (per rules.md)
        expected_sortant = (
            en_cours_montant
            + repris_montant
            + declares_montant
            - reglements_montant
            - rejet_montant
            + reevaluation_pos
            - reevaluation_neg
        )
        verif_diff = abs(expected_sortant - reserves_montant)
        verif_ok = verif_diff <= 1.0  # tolerance in case of rounding

        row_dict: dict[str, Any] = {
            "exercice": yr,
            "en_cours_nbre": en_cours_nbre,
            "en_cours_montant": en_cours_montant,
            "repris_nbre": repris_nbre,
            "repris_montant": repris_montant,
            "declares_nbre": declares_nbre,
            "declares_montant": declares_montant,
            "reglements_nbre": reglements_nbre,
            "reglements_montant": reglements_montant,
            "rejet_nbre": rejet_nbre,
            "rejet_montant": rejet_montant,
            "reevaluation_pos": reevaluation_pos,
            "reevaluation_neg": reevaluation_neg,
            "reserves_nbre": reserves_nbre,
            "reserves_montant": reserves_montant,
            "reserves_montant_old_signed": reserves_montant_old_signed,
            "reserves_montant_delta_new_minus_old_signed": reserves_montant_delta_new_minus_old_signed,
            "verif_diff": verif_diff,
            "verif_ok": verif_ok,
        }
        table.append(row_dict)
        prev_sortant_nbre = reserves_nbre
        prev_sortant_montant = reserves_montant

    all_ok = all(row["verif_ok"] for row in table)
    total_reserves_new = sum(float(row["reserves_montant"]) for row in table)
    total_reserves_old_signed = sum(float(row["reserves_montant_old_signed"]) for row in table)
    return {
        "generated_at": utc_isoformat(),
        "scope_mode": "portfolio",
        "sap_method": "method_b_workbook_ac",
        "sap_legacy_reference_method": "signed_outstanding_pre_clipping",
        "years": table,
        "total_reserves": prev_sortant_montant,
        "reserve_totals_comparison": {
            "total_reserves_new": total_reserves_new,
            "total_reserves_old_signed": total_reserves_old_signed,
            "delta_new_minus_old_signed": total_reserves_new - total_reserves_old_signed,
        },
        "reserves_delta_by_year": [
            {
                "exercice": row["exercice"],
                "delta_new_minus_old_signed": row["reserves_montant_delta_new_minus_old_signed"],
            }
            for row in table
        ],
        "all_years_balanced": all_ok,
        "source": "level3_bilan_sinistres",
    }


def list_audit_events(settings: BackendSettings, limit: int = 100) -> list[dict[str, Any]]:
    """Return audit events in reverse chronological order."""

    with get_connection(settings.db_path) as connection:
        rows = connection.execute(
            """
            SELECT event_id, actor_user_id, action, target_type, target_id, occurred_at,
                   ip_address, user_agent, payload_json, previous_event_hash, event_hash
            FROM audit_event
            ORDER BY occurred_at DESC, event_id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [
        {
            **_dict(row),
            "payload": json.loads(row["payload_json"]),
        }
        for row in rows
    ]


def get_audit_event(settings: BackendSettings, event_id: str) -> dict[str, Any]:
    """Return one audit event by ID."""

    with get_connection(settings.db_path) as connection:
        row = connection.execute(
            """
            SELECT event_id, actor_user_id, action, target_type, target_id, occurred_at,
                   ip_address, user_agent, payload_json, previous_event_hash, event_hash
            FROM audit_event
            WHERE event_id = ?
            """,
            (event_id,),
        ).fetchone()
        if row is None:
            raise BackendError(f"Unknown audit event: {event_id}", status_code=404)
    return {**_dict(row), "payload": json.loads(row["payload_json"])}  # type: ignore[arg-type]
