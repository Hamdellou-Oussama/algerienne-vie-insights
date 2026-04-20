from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
import pandas as pd


@dataclass
class TriangleData:
    incremental: pd.DataFrame
    cumulative_observed: pd.DataFrame
    factors: Dict[int, float]
    projected_cumulative: pd.DataFrame
    diagonal: Dict[int, float]
    ultimate: Dict[int, float]
    ibnr: Dict[int, float]


def find_ibnr_workbook(root: Path) -> Path:
    matches = sorted(root.glob("data/level 02-*DATA IBNR.xlsx"))
    if not matches:
        raise FileNotFoundError("Could not find level 02 IBNR workbook under data/")
    return matches[0]


def _pick_columns(df: pd.DataFrame) -> Tuple[str, str, str]:
    cols = [str(c) if c is not None else "" for c in df.columns]
    low = [c.lower() for c in cols]

    ay_idx = next(i for i, c in enumerate(low) if "ann" in c and "sinistre" in c)
    dy_idx = next(i for i, c in enumerate(low) if "ann" in c and "claration" in c)
    mt_idx = next(i for i, c in enumerate(low) if "montant" in c)
    return cols[ay_idx], cols[dy_idx], cols[mt_idx]


def load_base_ade_incremental(
    workbook_path: Path,
    accident_years: List[int] | None = None,
    dev_max: int = 3,
) -> pd.DataFrame:
    if accident_years is None:
        accident_years = [2022, 2023, 2024, 2025]

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb["base ADE"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    df = pd.DataFrame(rows, columns=[ws.cell(1, c).value for c in range(1, ws.max_column + 1)])

    col_ay, col_dy, col_mt = _pick_columns(df)
    work = df[[col_ay, col_dy, col_mt]].dropna().copy()
    work.columns = ["AY", "DY", "amount"]
    work = work[work["AY"].isin(accident_years)].copy()
    work["AY"] = work["AY"].astype(int)
    work["DY"] = work["DY"].astype(int)
    work["dev"] = (work["DY"] - work["AY"]).astype(int)
    work = work[(work["dev"] >= 0) & (work["dev"] <= dev_max)]

    inc = (
        work.groupby(["AY", "dev"]) ["amount"]
        .sum()
        .unstack("dev")
        .reindex(index=accident_years, columns=range(dev_max + 1))
    )
    return inc


def observed_cell(ay: int, dev: int, valuation_year: int = 2025) -> bool:
    return (valuation_year - ay) >= dev


def to_cumulative_observed(inc: pd.DataFrame, valuation_year: int = 2025) -> pd.DataFrame:
    dev_max = int(inc.columns.max())
    cum = inc.copy()
    for ay in inc.index:
        running = 0.0
        for d in range(dev_max + 1):
            if observed_cell(int(ay), d, valuation_year=valuation_year):
                running += float(inc.loc[ay, d])
                cum.loc[ay, d] = running
            else:
                cum.loc[ay, d] = pd.NA
    return cum


def chain_ladder_factors(cum_obs: pd.DataFrame) -> Dict[int, float]:
    # Matches workbook scope exactly for AY 2022..2025 and dev 0..3.
    factors: Dict[int, float] = {}
    factors[1] = float(cum_obs.loc[[2022, 2023, 2024], 1].sum() / cum_obs.loc[[2022, 2023, 2024], 0].sum())
    factors[2] = float(cum_obs.loc[[2022, 2023], 2].sum() / cum_obs.loc[[2022, 2023], 1].sum())
    factors[3] = float(cum_obs.loc[[2022], 3].sum() / cum_obs.loc[[2022], 2].sum())
    return factors


def project_with_factors(cum_obs: pd.DataFrame, factors: Dict[int, float]) -> pd.DataFrame:
    dev_max = int(cum_obs.columns.max())
    proj = cum_obs.copy()
    for ay in proj.index:
        for d in range(1, dev_max + 1):
            if pd.isna(proj.loc[ay, d]):
                proj.loc[ay, d] = float(proj.loc[ay, d - 1]) * factors[d]
    return proj


def diagonal_ultimate_ibnr(cum_obs: pd.DataFrame, proj: pd.DataFrame, valuation_year: int = 2025):
    dev_max = int(proj.columns.max())
    diagonal: Dict[int, float] = {}
    ultimate: Dict[int, float] = {}
    ibnr: Dict[int, float] = {}
    for ay in proj.index:
        d_obs = valuation_year - int(ay)
        diagonal[int(ay)] = float(cum_obs.loc[ay, d_obs])
        ultimate[int(ay)] = float(proj.loc[ay, dev_max])
        ibnr[int(ay)] = float(ultimate[int(ay)] - diagonal[int(ay)])
    return diagonal, ultimate, ibnr


def build_chain_ladder_data(workbook_path: Path) -> TriangleData:
    inc = load_base_ade_incremental(workbook_path)
    cum_obs = to_cumulative_observed(inc)
    factors = chain_ladder_factors(cum_obs)
    proj = project_with_factors(cum_obs, factors)
    diagonal, ultimate, ibnr = diagonal_ultimate_ibnr(cum_obs, proj)
    return TriangleData(
        incremental=inc,
        cumulative_observed=cum_obs,
        factors=factors,
        projected_cumulative=proj,
        diagonal=diagonal,
        ultimate=ultimate,
        ibnr=ibnr,
    )
