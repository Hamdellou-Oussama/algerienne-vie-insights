from __future__ import annotations

from pathlib import Path
import json
import py_compile
import subprocess
import sys

import openpyxl
import pandas as pd

from common_triangle import build_chain_ladder_data, find_ibnr_workbook, load_base_ade_incremental

METHOD_SCRIPTS = [
    "method_chain_ladder.py",
    "method_mack_chain_ladder.py",
    "method_bornhuetter_ferguson.py",
    "method_benktander.py",
    "method_bootstrap_odp.py",
    "method_cape_cod.py",
    "method_munich_chain_ladder.py",
    "method_glm_reserving.py",
]


def check_feature_names(workbook: Path) -> dict:
    inc = load_base_ade_incremental(workbook)
    return {
        "status": "ok",
        "shape": [int(inc.shape[0]), int(inc.shape[1])],
        "columns": [int(c) for c in inc.columns],
        "ays": [int(a) for a in inc.index],
    }


def check_formula_alignment(workbook: Path) -> dict:
    cl = build_chain_ladder_data(workbook)
    wb = openpyxl.load_workbook(workbook, data_only=True)
    ws = wb["calcule IBNR"]
    f_excel = {1: ws.cell(29, 3).value, 2: ws.cell(29, 4).value, 3: ws.cell(29, 5).value}
    i_excel = {
        2022: ws.cell(33, 9).value,
        2023: ws.cell(34, 9).value,
        2024: ws.cell(35, 9).value,
        2025: ws.cell(36, 9).value,
    }
    total_excel = ws.cell(37, 9).value

    f_diff = {k: float(cl.factors[k] - f_excel[k]) for k in [1, 2, 3]}
    i_diff = {k: float(cl.ibnr[k] - i_excel[k]) for k in [2022, 2023, 2024, 2025]}
    total_diff = float(sum(cl.ibnr.values()) - total_excel)

    return {
        "status": "ok",
        "factor_diff": f_diff,
        "ibnr_diff": i_diff,
        "total_diff": total_diff,
        "formula_check_pass": abs(total_diff) < 1e-2 and max(abs(v) for v in i_diff.values()) < 1e-2,
    }


def run_script(script_path: Path) -> dict:
    proc = subprocess.run([sys.executable, str(script_path)], capture_output=True, text=True)
    return {
        "returncode": proc.returncode,
        "stdout_tail": proc.stdout[-2500:],
        "stderr_tail": proc.stderr[-2500:],
        "status": "ok" if proc.returncode == 0 else "failed",
    }


def main() -> None:
    here = Path(__file__).resolve().parent
    root = here.parents[1]
    out_dir = here / "outputs"
    out_dir.mkdir(parents=True, exist_ok=True)

    workbook = find_ibnr_workbook(root)

    verification = {
        "check_1_syntax": {},
        "check_2_feature_names": {},
        "check_3_formulas": {},
        "check_4_runtime": {},
    }

    # 1) Syntax
    for script in METHOD_SCRIPTS:
        path = here / script
        try:
            py_compile.compile(str(path), doraise=True)
            verification["check_1_syntax"][script] = "ok"
        except Exception as e:
            verification["check_1_syntax"][script] = f"failed: {e}"

    # 2) Feature names / triangle extraction viability
    verification["check_2_feature_names"] = check_feature_names(workbook)

    # 3) Formula reconciliation against workbook for chain-ladder core
    verification["check_3_formulas"] = check_formula_alignment(workbook)

    # 4) Runtime per method
    for script in METHOD_SCRIPTS:
        verification["check_4_runtime"][script] = run_script(here / script)

    with open(out_dir / "verification_4checks.json", "w", encoding="utf-8") as f:
        json.dump(verification, f, indent=2)

    print("Wrote:", out_dir / "verification_4checks.json")
    print("Formula check pass:", verification["check_3_formulas"]["formula_check_pass"])


if __name__ == "__main__":
    main()
